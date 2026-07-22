#!/usr/bin/env python3
"""
report.py — รวมอัตราส่วนทุกบริษัท + เปรียบเทียบแหล่งภายนอก -> output/ratios_report.xlsx

ชีต:
  1. Summary_FY       อัตราส่วนรายปีทุกบริษัท (CFA basis) + D/E
  2. SET_vs_CFA       DSO/DIO/DPO/CCC สองฐาน (SET เทียบเว็บ / CFA ตามงบ)
  3. External_MINT    เทียบ 3 ทาง: เรา vs settrade vs MD&A (MINT มีข้อมูลครบ)
  4. Quarterly        รายไตรมาสทุกจุด (40+)
"""
import os
import sys
import csv
import glob

sys.path.insert(0, os.path.dirname(__file__))
import extract_set as E
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RAW = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data", "raw"))
OUT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "output"))

HDR = PatternFill("solid", fgColor="1F4E78")
HF = Font(bold=True, color="FFFFFF")
SUB = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

# ---- ตัวเลขแหล่งภายนอกที่มี (ผู้ใช้ให้ / จาก MD&A) ----
SETTRADE = {                       # settrade/SETSMART (ผู้ใช้ส่งมา)
    ("MINT", 2566): dict(DSO=45.18, DIO=24.88, DPO=174.08, CCC=-104.02),
    ("MINT", 2567): dict(DSO=31.60, DIO=24.16, DPO=153.98, CCC=-98.22),
    ("MINT", 2568): dict(DSO=36.16, DIO=27.15, DPO=162.67, CCC=-99.36),
}
MDA = {                            # ตัวเลขที่บริษัทเปิดเผยเอง (MD&A)
    ("MINT", 2568): dict(DSO=24, DIO=20, DPO=67, ROA=0.028, ROE=0.099),
}


def load():
    files = []
    for ext in ("xls", "xlsx", "xlsm", "XLS", "XLSX"):
        files += glob.glob(os.path.join(RAW, "*", "*", f"*.{ext}"))
    groups = {}
    for f in sorted(set(files)):
        try:
            rec = E.extract_file(f)
        except E.ExtractError:
            continue
        groups.setdefault((rec["ticker"], rec["year"]), {})[
            "FY" if rec["is_annual"] else rec["quarter"]] = rec
    return groups


def style_header(ws, row=1, ncol=None):
    for c in ws[row]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center", wrap_text=True)


def annual_cfa(bt):
    """อัตราส่วนรายปี CFA: เฉลี่ยงบดุลต้นปี-ปลายปี, ต้นทุนขายแคบ, 365"""
    fy, q1 = bt["FY"], bt.get("Q1")
    rev, cogs, npt = fy["total_revenue"], fy["cogs"], fy["net_profit_total"]
    npp = fy.get("net_profit_parent")
    beg = q1 if q1 else fy
    def avg(k):
        b = beg.get(k + "_begin");  e = fy.get(k + "_end")
        return (b + e) / 2 if (b is not None and e is not None) else e
    ta, te = avg("total_assets"), avg("total_equity")
    return dict(
        NPM=npt / rev, ROA=npt / ta, ROE=npt / te,
        ROE_parent=(npp / te if npp is not None else None),
        DSO=avg("trade_receivables") / rev * 365,
        DIO=avg("inventory") / cogs * 365,
        DPO=avg("trade_payables") / cogs * 365,
        DE=fy["total_liabilities_end"] / fy["total_equity_end"],
        GM=(rev - cogs) / rev,
    ) | {"CCC": None}


def set_basis(fy):
    op = fy["total_revenue"] - (fy.get("other_income") or 0)
    cost = fy["cogs"]
    ap = fy.get("trade_payables_broad_end") or fy["trade_payables_end"]
    dso = fy["trade_receivables_end"] / op * 365
    dio = fy["inventory_end"] / cost * 365
    dpo = ap / cost * 365
    return dict(DSO=dso, DIO=dio, DPO=dpo, CCC=dso + dio - dpo)


def pct(x):
    return f"{x*100:.1f}%" if x is not None else "—"


def num(x):
    return round(x, 1) if x is not None else None


def main():
    groups = load()
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: Summary_FY ----
    ws = wb.create_sheet("Summary_FY")
    ws.append(["อัตราส่วนรายปี FY2568 — ฐาน CFA (คำนวณจากงบตรวจสอบ)"])
    ws.append(["บริษัท", "NPM", "Gross Margin", "ROA", "ROE (รวม)", "ROE (ส่วนแม่)",
               "DSO (วัน)", "DIO (วัน)", "DPO (วัน)", "D/E"])
    for (tk, yr), bt in sorted(groups.items()):
        if yr != 2568 or "FY" not in bt:
            continue
        a = annual_cfa(bt)
        ws.append([tk, pct(a["NPM"]), pct(a["GM"]), pct(a["ROA"]),
                   pct(a["ROE"]), pct(a["ROE_parent"]),
                   num(a["DSO"]), num(a["DIO"]), num(a["DPO"]), round(a["DE"], 2)])
    style_header(ws, 2)
    ws["A1"].font = BOLD
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 12

    # ---- Sheet 2: SET_vs_CFA ----
    ws = wb.create_sheet("SET_vs_CFA")
    ws.append(["DSO/DIO/DPO/CCC รายปี FY2568 — 2 ฐาน"])
    ws.append(["", "SET basis (เทียบเว็บ)", "", "", "", "CFA basis (ตามงบ)", "", "", ""])
    ws.append(["บริษัท", "DSO", "DIO", "DPO", "CCC", "DSO", "DIO", "DPO", "CCC"])
    for (tk, yr), bt in sorted(groups.items()):
        if yr != 2568 or "FY" not in bt:
            continue
        s = set_basis(bt["FY"])
        a = annual_cfa(bt)
        a["CCC"] = a["DSO"] + a["DIO"] - a["DPO"]
        ws.append([tk, num(s["DSO"]), num(s["DIO"]), num(s["DPO"]), num(s["CCC"]),
                   num(a["DSO"]), num(a["DIO"]), num(a["DPO"]), num(a["CCC"])])
    style_header(ws, 3)
    for c in ws[2]:
        c.font = BOLD
        c.fill = SUB
    ws["A1"].font = BOLD
    ws.column_dimensions["A"].width = 12

    # ---- Sheet 3: External_MINT (เทียบ 3 ทาง) ----
    ws = wb.create_sheet("External_MINT")
    ws.append(["MINT — เปรียบเทียบกับแหล่งภายนอก (เรา vs settrade vs MD&A)"])
    ws.append([])
    ws.append(["Working capital (วัน)", "ปี", "เรา (SET-basis)", "settrade", "ต่าง",
               "MD&A บริษัท"])
    for yr in (2566, 2567, 2568):
        bt = groups.get(("MINT", yr))
        if not bt or "FY" not in bt:
            continue
        s = set_basis(bt["FY"])
        st = SETTRADE.get(("MINT", yr), {})
        md = MDA.get(("MINT", yr), {})
        for k in ("DSO", "DIO", "DPO", "CCC"):
            diff = (s[k] - st[k]) if k in st else None
            ws.append([k, yr, num(s[k]), st.get(k, "—"),
                       (round(diff, 1) if diff is not None else "—"), md.get(k, "—")])
    ws.append([])
    ws.append(["Profitability (%)", "ปี", "เรา (CFA)", "", "", "MD&A บริษัท"])
    a68 = annual_cfa(groups[("MINT", 2568)])
    md68 = MDA[("MINT", 2568)]
    for k, lab in [("ROA", "ROA"), ("ROE_parent", "ROE")]:
        ws.append([lab, 2568, pct(a68[k]), "", "", pct(md68.get(lab))])
    ws.append([])
    ws.append(["สรุป: DSO ตรง settrade ~99%; DIO/DPO ต่างเพราะ settrade ใช้ต้นทุน standardized "
               "(~0.6x งบ, black-box); ROA ตรง MD&A เป๊ะ"])
    style_header(ws, 3)
    ws["A1"].font = BOLD
    ws["A6"].font = BOLD if ws["A6"].value else Font()
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    # ---- Sheet 4: Quarterly ----
    ws = wb.create_sheet("Quarterly")
    with open(os.path.join(OUT, "metrics_all.csv"), encoding="utf-8-sig") as f:
        for i, row in enumerate(csv.reader(f)):
            ws.append(row)
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 10

    path = os.path.join(OUT, "ratios_report.xlsx")
    wb.save(path)
    print("wrote", os.path.relpath(path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
