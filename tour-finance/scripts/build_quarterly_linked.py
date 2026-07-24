#!/usr/bin/env python3
"""
build_quarterly_linked.py — สมุดงาน Excel "สูตรสด" 5 ปี × 4 ไตรมาส ต่อบริษัท

ต่างจาก build_5yr.py (ต้นแบบที่ดึงกำไรสุทธิจาก MD&A แบบสะสม): เวอร์ชันนี้ใช้
"งบกำไรขาดทุน 3 เดือนจริง" ที่ extractor ดึงมา จึงไม่ต้องพึ่ง MD&A และเติมข้อมูล
จริงลงช่อง input ให้อัตโนมัติเท่าที่มีไฟล์ ส่วนไตรมาสที่ยังไม่มีไฟล์ = ช่องว่าง
(เหลือง) ที่รอเติม — เติมเมื่อไร สูตรในชีต Ratios คำนวณเองทันที

3 ชีตข้อมูล + Ratios + Notes:
  IS_Flow      : รายได้/รายได้อื่น/ต้นทุนขาย/กำไรสุทธิ รายไตรมาส (3 เดือน)
                 คอลัมน์ Q4 = =FY − Q1 − Q2 − Q3 (สูตรสด)
  BS_Timeline  : ยอดคงเหลือ 21 จุดเวลา (ต้นปี 2563 + 20 สิ้นไตรมาส)
  Ratios       : 7 ตัวชี้วัด + GM + D/E × 20 ไตรมาส (=สูตรสด อ้าง IS_Flow/BS_Timeline)
                 งบดุลเฉลี่ยแบบ chain: ไตรมาส n ใช้ (ปลายไตรมาสก่อน + ปลายไตรมาสนี้)/2

ใช้:
    python3 build_quarterly_linked.py <TICKER>     # output/<TICKER>_quarterly_linked.xlsx
    python3 build_quarterly_linked.py --all
"""
import os
import sys
import glob

sys.path.insert(0, os.path.dirname(__file__))
import extract_set as E
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as GL

TICKERS = ["ASIA", "CENTEL", "DUSIT", "ERW", "MANRIN", "MINT", "OHTL", "SHANG", "SHR", "VRANDA"]
YEARS = [2564, 2565, 2566, 2567, 2568]
QTRS = ["Q1", "Q2", "Q3", "Q4"]
DAYS = [90, 91, 92, 92]
MONTHS = ["Mar", "Jun", "Sep", "Dec"]

RAW = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data", "raw"))
OUT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "output"))

HDR = PatternFill("solid", fgColor="1F4E78")
HF = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")     # เหลือง = ช่องเติมข้อมูล
FORMULA_FILL = PatternFill("solid", fgColor="E2EFDA")   # เขียวอ่อน = สูตร
DERIVED_FILL = PatternFill("solid", fgColor="FCE4D6")   # ส้มอ่อน = Q4 (สูตรลบ)
BOLD = Font(bold=True)

# แถวในชีต IS_Flow
IS_ROWS = ["รายได้รวม", "รายได้อื่น", "ต้นทุนขาย (จริง)", "กำไรสุทธิ (รวม)", "กำไรสุทธิ (ส่วนแม่)"]
IS_KEYS = ["total_revenue", "other_income", "cogs", "net_profit_total", "net_profit_parent"]
# แถวในชีต BS_Timeline
BS_ROWS = ["ลูกหนี้การค้า", "สินค้าคงเหลือ", "เจ้าหนี้การค้า", "สินทรัพย์รวม",
           "ส่วนของเจ้าของรวม", "หนี้สินรวม"]
BS_KEYS = ["trade_receivables", "inventory", "trade_payables", "total_assets",
           "total_equity", "total_liabilities"]


def style_header(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center", wrap_text=True)


def load_company(ticker):
    """คืน {year: {tag: rec}} ของบริษัทเดียว (tag = Q1/Q2/Q3/FY)"""
    files = []
    for ext in ("xls", "xlsx", "xlsm", "XLS", "XLSX"):
        files += glob.glob(os.path.join(RAW, ticker, "*", f"*.{ext}"))
    data = {}
    for f in sorted(set(files)):
        try:
            rec = E.extract_file(f)
        except E.ExtractError:
            continue
        if rec["ticker"] != ticker:
            continue
        tag = "FY" if rec["is_annual"] else rec["quarter"]
        data.setdefault(rec["year"], {})[tag] = rec
    return data


def bs_col(year, month_idx):
    """คอลัมน์ใน BS_Timeline ของ (ปี, เดือนสิ้นไตรมาส); slot0(col2)=ต้นปี 2564"""
    return 3 + (year - YEARS[0]) * 4 + month_idx


def prev_dec_col(year):
    """คอลัมน์ 'ปลายปีก่อน' (= ต้นงวดของปีนี้)"""
    if year == YEARS[0]:
        return 2                          # slot0 = 31-Dec-2563
    return bs_col(year - 1, 3)            # Dec ของปีก่อน


# ---------- Sheet: IS_Flow ----------
def build_is_flow(wb, data):
    ws = wb.create_sheet("IS_Flow")
    hdr = ["รายการ (ล้านบาท, 3 เดือน)"]
    for y in YEARS:
        hdr += [f"Q1/{y}", f"Q2/{y}", f"Q3/{y}", f"Q4/{y}", f"FY{y}"]
    ws.append(hdr)
    for label in IS_ROWS:
        ws.append([label])
    # เติมค่า 3 เดือนจริง — ดึงผ่าน compute_group เพื่อ de-cumulate งบสะสม (MANRIN)
    # และ derive Q4 ให้ถูกต้องทุกรูปแบบการยื่นงบ
    for yi, y in enumerate(YEARS):
        base = 2 + yi * 5                  # col Q1
        recs = data.get(y, {})
        qflow = {rec["quarter"]: rec for rec, _ in E.compute_group(recs)}
        fy_rec = recs.get("FY")
        cumulative = E._cumulative_year(recs)
        for ri, key in enumerate(IS_KEYS, start=2):
            cq1, cq2, cq3, cq4, cfy = base, base + 1, base + 2, base + 3, base + 4
            # ผูก Q4=FY−Q1−Q2−Q3 ได้ก็ต่อเมื่อ FY ของ "รายการนี้" เป็นตัวเลขจริง (กัน circular)
            fy_val_ok = fy_rec is not None and isinstance(fy_rec.get(key), (int, float))
            std_q4 = fy_val_ok and not cumulative
            for qi, col in enumerate((cq1, cq2, cq3)):
                rec = qflow.get(("Q1", "Q2", "Q3")[qi])
                v = rec.get(key) if rec else None
                ws.cell(ri, col).value = round(v, 3) if isinstance(v, (int, float)) else None
                if v is None:
                    ws.cell(ri, col).fill = INPUT_FILL
            if std_q4:                     # Q4 = FY − Q1 − Q2 − Q3 (สูตรสด)
                ws.cell(ri, cq4).value = (f"={GL(cfy)}{ri}-{GL(cq1)}{ri}"
                                          f"-{GL(cq2)}{ri}-{GL(cq3)}{ri}")
                ws.cell(ri, cq4).fill = DERIVED_FILL
            else:                          # งบสะสม/ไม่มี FY ของรายการนี้ -> ใส่ค่า Q4 จริง
                r4 = qflow.get("Q4")
                v4 = r4.get(key) if r4 else None
                ws.cell(ri, cq4).value = round(v4, 3) if isinstance(v4, (int, float)) else None
                if v4 is None:
                    ws.cell(ri, cq4).fill = INPUT_FILL
            if fy_val_ok:
                ws.cell(ri, cfy).value = round(fy_rec[key], 3)
            else:                          # FY = ผลรวม 4 ไตรมาส (Q4 เป็นค่า -> ไม่วน)
                ws.cell(ri, cfy).value = (f"={GL(cq1)}{ri}+{GL(cq2)}{ri}"
                                          f"+{GL(cq3)}{ri}+{GL(cq4)}{ri}")
                ws.cell(ri, cfy).fill = DERIVED_FILL
    for r in range(2, 2 + len(IS_ROWS)):
        ws.cell(r, 1).font = BOLD
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 26
    return ws


# ---------- Sheet: BS_Timeline ----------
def build_bs_timeline(wb, data):
    ws = wb.create_sheet("BS_Timeline")
    hdr = ["รายการ (ยอดคงเหลือ, ล้านบาท)", f"31-Dec-{YEARS[0]-1}"]
    for y in YEARS:
        hdr += [f"31-Mar-{y}", f"30-Jun-{y}", f"30-Sep-{y}", f"31-Dec-{y}"]
    ws.append(hdr)
    for label in BS_ROWS:
        ws.append([label] + [None] * 21)

    # เติมยอดปลายไตรมาสจากไฟล์ที่มี
    for ri, key in enumerate(BS_KEYS, start=2):
        for y in YEARS:
            recs = data.get(y, {})
            # Mar/Jun/Sep จากไฟล์ Q1/Q2/Q3 (ยอดปลายงวด)
            for tag, mi in [("Q1", 0), ("Q2", 1), ("Q3", 2)]:
                rec = recs.get(tag)
                if rec and rec.get(key + "_end") is not None:
                    ws.cell(ri, bs_col(y, mi)).value = round(rec[key + "_end"], 3)
            # Dec จากไฟล์ FY (ยอดปลายปี)
            fy = recs.get("FY")
            if fy and fy.get(key + "_end") is not None:
                ws.cell(ri, bs_col(y, 3)).value = round(fy[key + "_end"], 3)
            # ต้นงวด: จากยอด begin ของไฟล์ Q1 (หรือ FY) -> ปลายปีก่อน
            src = recs.get("Q1") or recs.get("FY")
            if src and src.get(key + "_begin") is not None:
                pc = prev_dec_col(y)
                if ws.cell(ri, pc).value is None:
                    ws.cell(ri, pc).value = round(src[key + "_begin"], 3)
        # ช่องที่ยังว่าง = เหลือง (รอเติม)
        for c in range(2, 23):
            if ws.cell(ri, c).value is None:
                ws.cell(ri, c).fill = INPUT_FILL
        ws.cell(ri, 1).font = BOLD
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 22
    return ws


# ---------- Sheet: Ratios (สูตรสด) ----------
def build_ratios(wb, ticker):
    ws = wb.create_sheet("Ratios")
    ws.append(["ตัวชี้วัด"] + [f"{q}/{y}" for y in YEARS for q in QTRS])
    labels = ["รายได้รวม (ล้านบาท)", "ต้นทุนขาย (ล้านบาท)", "กำไรสุทธิ (ล้านบาท)",
              "NPM (%)", "Gross Margin (%)", "DSO (วัน)", "DIO (วัน)", "DPO (วัน)",
              "CCC (วัน)", "ROA รายไตรมาส (%)", "ROE รายไตรมาส (%)", "D/E (เท่า)"]
    for lab in labels:
        ws.append([lab])
    # แถว: rev2 cogs3 np4 npm5 gm6 dso7 dio8 dpo9 ccc10 roa11 roe12 de13
    pct_rows = {5, 6, 10, 11}
    for q in range(20):
        yi, qi = q // 4, q % 4
        col = 2 + q
        Q = GL(col)
        d = DAYS[qi]
        ic = 2 + yi * 5 + qi                 # คอลัมน์ IS_Flow ของไตรมาสนี้
        I = GL(ic)
        rev = f"IS_Flow!{I}2"
        oth = f"IS_Flow!{I}3"
        cogs = f"IS_Flow!{I}4"
        np_ = f"IS_Flow!{I}5"
        bb, be = GL(2 + q), GL(3 + q)        # begin/end col ใน BS_Timeline (chain)

        def avg(row):
            return f"((BS_Timeline!{bb}{row}+BS_Timeline!{be}{row})/2)"

        ws.cell(2, col).value = f"={rev}"
        ws.cell(3, col).value = f"={cogs}"
        ws.cell(4, col).value = f"={np_}"
        ws.cell(5, col).value = f'=IFERROR({np_}/{rev},"")'
        ws.cell(6, col).value = f'=IFERROR(({rev}-{oth}-{cogs})/({rev}-{oth}),"")'
        ws.cell(7, col).value = f'=IFERROR({avg(2)}/{rev}*{d},"")'
        ws.cell(8, col).value = f'=IFERROR({avg(3)}/{cogs}*{d},"")'
        ws.cell(9, col).value = f'=IFERROR({avg(4)}/{cogs}*{d},"")'
        ws.cell(10, col).value = f'=IFERROR({Q}7+{Q}8-{Q}9,"")'
        ws.cell(11, col).value = f'=IFERROR({np_}/{avg(5)},"")'
        ws.cell(12, col).value = f'=IFERROR({np_}/{avg(6)},"")'
        ws.cell(13, col).value = f'=IFERROR(BS_Timeline!{be}7/BS_Timeline!{be}6,"")'
        for r in range(2, 14):
            ws.cell(r, col).fill = FORMULA_FILL
            ws.cell(r, col).number_format = "0.0%" if r in pct_rows else "0.0"
    for r in range(2, 2 + len(labels)):
        ws.cell(r, 1).font = BOLD
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 22
    for q in range(20):
        ws.column_dimensions[GL(2 + q)].width = 10
    return ws


def build_notes(wb, ticker, data):
    ws = wb.create_sheet("Notes")
    have = ", ".join(f"{tag}/{y}" for y in YEARS for tag in ("Q1", "Q2", "Q3", "FY")
                     if tag in data.get(y, {}))
    lines = [
        f"{ticker} — สมุดงานสูตรสด 5 ปี × 4 ไตรมาส (ฐาน CFA)",
        "",
        "สีของช่อง:  เหลือง = ช่องเติมตัวเลขดิบ (แก้ได้)   เขียว = สูตรสด   ส้ม = Q4 (=FY−Q1−Q2−Q3)",
        "",
        "แหล่งข้อมูล: งบการเงินที่ตลาดหลักทรัพย์เผยแพร่ (งบกำไรขาดทุน 3 เดือน + งบฐานะการเงิน)",
        f"ไตรมาสที่มีข้อมูลจริงแล้ว: {have or '(ยังไม่มี)'}",
        "ไตรมาสอื่น = ช่องเหลืองรอเติม; เติมใน IS_Flow / BS_Timeline แล้ว Ratios คำนวณเอง",
        "",
        "สูตร (ฐาน CFA, ค่าเฉลี่ยงบดุลแบบ chain ปลายไตรมาสก่อน→ปลายไตรมาสนี้, วันตามจริง 90/91/92/92):",
        "  NPM = กำไรสุทธิ / รายได้รวม",
        "  Gross Margin = (รายได้ดำเนินงาน − ต้นทุนขาย) / รายได้ดำเนินงาน  [รายได้ดำเนินงาน = รายได้รวม − รายได้อื่น]",
        "  DSO = เฉลี่ยลูกหนี้การค้า / รายได้รวม × วัน",
        "  DIO = เฉลี่ยสินค้าคงเหลือ / ต้นทุนขาย × วัน",
        "  DPO = เฉลี่ยเจ้าหนี้การค้า / ต้นทุนขาย × วัน",
        "  CCC = DSO + DIO − DPO",
        "  ROA = กำไรสุทธิ / เฉลี่ยสินทรัพย์รวม (รายไตรมาส ไม่ annualize)",
        "  ROE = กำไรสุทธิ / เฉลี่ยส่วนของเจ้าของ (รายไตรมาส ไม่ annualize)",
        "  D/E = หนี้สินรวม / ส่วนของเจ้าของรวม (ยอดปลายไตรมาส)",
        "",
        "หมายเหตุ chain งบดุล: งบไตรมาส SET เทียบ 'ต้นปี' เสมอ — BS_Timeline จึงเก็บยอดปลายทุก",
        "ไตรมาสต่อเนื่อง เพื่อให้ค่าเฉลี่ยของไตรมาส n = (ปลายไตรมาส n−1 + ปลายไตรมาส n)/2 ถูกต้อง",
    ]
    for ln in lines:
        ws.append([ln])
    ws.column_dimensions["A"].width = 104
    ws["A1"].font = BOLD
    return ws


def build_company(ticker):
    data = load_company(ticker)
    wb = Workbook()
    wb.remove(wb.active)
    build_is_flow(wb, data)
    build_bs_timeline(wb, data)
    build_ratios(wb, ticker)
    build_notes(wb, ticker, data)
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, f"{ticker}_quarterly_linked.xlsx")
    wb.save(path)
    n = sum(1 for y in YEARS for tag in ("Q1", "Q2", "Q3", "FY") if tag in data.get(y, {}))
    return os.path.normpath(path), n


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    targets = TICKERS if argv[1] == "--all" else [argv[1].upper()]
    for t in targets:
        if t not in TICKERS:
            print(f"unknown ticker {t}; valid: {', '.join(TICKERS)}")
            return 1
        path, n = build_company(t)
        print(f"wrote {os.path.relpath(path)}  (ไฟล์จริง {n} งวด)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
