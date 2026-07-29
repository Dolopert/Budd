#!/usr/bin/env python3
"""
annual_avg_compare.py — เทียบค่าเฉลี่ยงบดุลรายปี 2 จุด vs 5 จุด (สำรองไว้ใช้เมื่อข้อมูลครบ)

ค่าเฉลี่ยงบดุลรายปีมี 2 นิยาม:
  2 จุด (ตำรา CFA default) : (ยอดต้นปี + ยอดปลายปี) / 2
  5 จุด (เหมาะธุรกิจฤดูกาล) : เฉลี่ย (ต้นปี, สิ้น Q1, สิ้น Q2, สิ้น Q3, ปลายปี)

ไฟล์นี้คำนวณ DSO/DIO/DPO/ROA/ROE รายปีทั้งสองแบบ + %ต่าง เพื่อดูว่าฤดูกาลมีผลจริงไหม
ต้องมีงบครบ Q1/Q2/Q3/FY ของปีนั้น (ถ้าไม่ครบ 5 จุดคำนวณไม่ได้ -> ข้าม)

ใช้:
    python3 annual_avg_compare.py [ปี]              # พิมพ์ตารางเทียบ (ค่าเริ่มต้น 2568)
    python3 annual_avg_compare.py [ปี] --xlsx       # เขียน output/annual_avg_compare_<ปี>.xlsx
"""
import os
import sys
import glob

sys.path.insert(0, os.path.dirname(__file__))
import extract_set as E

RAW = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data", "raw"))
OUT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "output"))
BS = ["trade_receivables", "inventory", "trade_payables", "total_assets", "total_equity"]


def load(year):
    files = []
    for ext in ("xls", "xlsx", "xlsm", "XLS", "XLSX"):
        files += glob.glob(os.path.join(RAW, "*", "*", f"*.{ext}"))
    groups = {}
    for f in sorted(set(files)):
        try:
            rec = E.extract_file(f)
        except E.ExtractError:
            continue
        if rec["year"] != year:
            continue
        tag = "FY" if rec["is_annual"] else rec["quarter"]
        groups.setdefault(rec["ticker"], {})[tag] = rec
    return groups


def points(bt, key):
    """คืน 5 จุดยอดงบดุลตลอดปี: [ต้นปี, สิ้นQ1, สิ้นQ2, สิ้นQ3, ปลายปี] (None ถ้าไม่ครบ)"""
    need = ("Q1", "Q2", "Q3", "FY")
    if any(t not in bt for t in need):
        return None
    q1, q2, q3, fy = (bt[t] for t in need)
    pts = [q1.get(key + "_begin"), q1.get(key + "_end"), q2.get(key + "_end"),
           q3.get(key + "_end"), fy.get(key + "_end")]
    return pts if all(p is not None for p in pts) else None


def ratios(bt, mode):
    """mode = '2pt' หรือ '5pt' -> dict อัตราส่วนรายปี (365 วัน)"""
    fy = bt["FY"]
    rev, cogs, npt = fy["total_revenue"], fy["cogs"], fy["net_profit_total"]
    avg = {}
    for key in BS:
        pts = points(bt, key)
        if pts is None:
            return None
        avg[key] = (pts[0] + pts[4]) / 2 if mode == "2pt" else sum(pts) / 5
    return dict(
        DSO=avg["trade_receivables"] / rev * 365,
        DIO=avg["inventory"] / cogs * 365,
        DPO=avg["trade_payables"] / cogs * 365,
        ROA=npt / avg["total_assets"],
        ROE=npt / avg["total_equity"],
    )


def main(argv):
    year = 2568
    want_xlsx = "--xlsx" in argv
    for a in argv[1:]:
        if a.isdigit():
            year = int(a)
    groups = load(year)
    rows = []
    for tk in sorted(groups):
        bt = groups[tk]
        if "FY" not in bt:
            continue
        r2, r5 = ratios(bt, "2pt"), ratios(bt, "5pt")
        if r2 is None or r5 is None:
            continue
        rows.append((tk, r2, r5))

    if not rows:
        print(f"ไม่มีบริษัทที่มีงบครบ Q1/Q2/Q3/FY ปี {year} (คำนวณ 5 จุดไม่ได้)")
        return 1

    metrics = ["DSO", "DIO", "DPO", "ROA", "ROE"]
    print(f"\n=== เทียบค่าเฉลี่ยงบดุลรายปี FY{year}: 2 จุด vs 5 จุด ===")
    hdr = f"{'บริษัท':8}" + "".join(f"{m+'(2)':>9}{m+'(5)':>9}{'Δ%':>7}" for m in metrics)
    print(hdr)
    is_pct = {"ROA", "ROE"}
    for tk, r2, r5 in rows:
        line = f"{tk:8}"
        for m in metrics:
            a, b = r2[m], r5[m]
            if m in is_pct:
                a, b = a * 100, b * 100
            d = (b - a) / a * 100 if a else 0
            line += f"{a:>9.1f}{b:>9.1f}{d:>6.0f}%"
        print(line)

    # ค่าเฉลี่ย |Δ%| ต่อ metric — บอกภาพรวมว่าฤดูกาลมีผลแค่ไหน
    print("\nΔ% เฉลี่ย (|ต่าง| ระหว่าง 2 จุด vs 5 จุด) ต่อ metric:")
    for m in metrics:
        diffs = []
        for _, r2, r5 in rows:
            a, b = r2[m], r5[m]
            if a:
                diffs.append(abs((b - a) / a * 100))
        print(f"  {m:4}: {sum(diffs)/len(diffs):5.1f}%  (สูงสุด {max(diffs):.0f}%)")

    if want_xlsx:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = f"AvgCompare_{year}"
        ws.append([f"เทียบค่าเฉลี่ยงบดุลรายปี FY{year}: 2 จุด vs 5 จุด (365 วัน)"])
        head = ["บริษัท"]
        for m in metrics:
            head += [f"{m} 2จุด", f"{m} 5จุด", f"{m} Δ%"]
        ws.append(head)
        for tk, r2, r5 in rows:
            row = [tk]
            for m in metrics:
                a, b = r2[m], r5[m]
                if m in is_pct:
                    a, b = a * 100, b * 100
                d = (b - a) / a * 100 if a else 0
                row += [round(a, 1), round(b, 1), round(d, 0)]
            ws.append(row)
        for c in ws[2]:
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws["A1"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 10
        os.makedirs(OUT, exist_ok=True)
        path = os.path.join(OUT, f"annual_avg_compare_{year}.xlsx")
        wb.save(path)
        print("\nwrote", os.path.relpath(path))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
