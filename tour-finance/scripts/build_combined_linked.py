#!/usr/bin/env python3
"""
build_combined_linked.py — รวมทุกบริษัทเป็นไฟล์ Excel เดียว (สูตรสดทั้งหมด)

output/TOUR_all_linked.xlsx — 6 ชีต:
  1. Summary_FY        อัตราส่วนรายปี FY2568 ทุกบริษัท (=สูตรสด อ้าง FY_Inputs)
  2. FY_Inputs         ตัวเลขดิบรายปี ทุกบริษัท (ช่องเหลือง แก้ได้)
  3. Quarterly_Ratios  7 ตัวชี้วัด+GM+D/E × 20 ไตรมาส ทุกบริษัท (=สูตรสด)
  4. IS_Flow           งบกำไรขาดทุน 3 เดือน ทุกบริษัท (Q4 = =FY−Q1−Q2−Q3)
  5. BS_Timeline       ยอดคงเหลือ 21 จุดเวลา ทุกบริษัท ( chain averaging)
  6. Notes             อธิบายสี + สูตร

เติมข้อมูลปีใดใน 3 ชีต input (FY_Inputs / IS_Flow / BS_Timeline) ชีตอัตราส่วน
คำนวณเองทันที (สูตรทั้งหมดผูกข้ามชีตในไฟล์เดียวกัน)

ใช้:  python3 build_combined_linked.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))
import report_linked as RL
import build_quarterly_linked as BQ
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as GL

TICKERS = BQ.TICKERS
YEARS = BQ.YEARS
QTRS = BQ.QTRS
DAYS = BQ.DAYS
OUT = BQ.OUT

HDR = PatternFill("solid", fgColor="1F4E78")
HF = Font(bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2EFDA")
DERIVED_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)

# แถวข้อมูลต่อบริษัทในแต่ละชีต (ไม่รวมแถว title)
IS_ROWS = BQ.IS_ROWS
IS_KEYS = BQ.IS_KEYS
BS_ROWS = BQ.BS_ROWS
BS_KEYS = BQ.BS_KEYS
IS_BLOCK = 1 + len(IS_ROWS)          # title + 5
BS_BLOCK = 1 + len(BS_ROWS)          # title + 6
RATIO_LABELS = ["รายได้รวม (ล้านบาท)", "ต้นทุนขาย (ล้านบาท)", "กำไรสุทธิ (ล้านบาท)",
                "NPM (%)", "Gross Margin (%)", "DSO (วัน)", "DIO (วัน)", "DPO (วัน)",
                "CCC (วัน)", "ROA รายไตรมาส (%)", "ROE รายไตรมาส (%)", "D/E (เท่า)"]
RATIO_BLOCK = 1 + len(RATIO_LABELS)  # title + 12


def style_header(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center", wrap_text=True)


def title_row(ws, r, text, ncol):
    cell = ws.cell(r, 1, text)
    cell.font = BOLD
    for c in range(1, ncol + 1):
        ws.cell(r, c).fill = TITLE_FILL


# ===== row offsets per company =====
def is_title(i):   return 2 + i * IS_BLOCK
def is_row(i, k):  return is_title(i) + 1 + k          # k = index in IS_KEYS
def bs_title(i):   return 2 + i * BS_BLOCK
def bs_row(i, k):  return bs_title(i) + 1 + k
def rt_title(i):   return 2 + i * RATIO_BLOCK
def rt_row(i, k):  return rt_title(i) + 1 + k


# ---------- Sheet: IS_Flow (ทุกบริษัท) ----------
def build_is_flow(wb, alldata):
    ws = wb.create_sheet("IS_Flow")
    hdr = ["รายการ (ล้านบาท, 3 เดือน)"]
    for y in YEARS:
        hdr += [f"Q1/{y}", f"Q2/{y}", f"Q3/{y}", f"Q4/{y}", f"FY{y}"]
    ncol = len(hdr)
    ws.append(hdr)
    import extract_set as E
    for i, t in enumerate(TICKERS):
        data = alldata[t]
        title_row(ws, is_title(i), f"■ {t}", ncol)
        # เตรียมค่า 3 เดือนจริงต่อปี (de-cumulate งบสะสม + derive Q4 ผ่าน compute_group)
        peryear = {}
        for y in YEARS:
            recs = data.get(y, {})
            qflow = {rec["quarter"]: rec for rec, _ in E.compute_group(recs)}
            peryear[y] = (qflow, recs.get("FY"), E._cumulative_year(recs))
        for k, (label, key) in enumerate(zip(IS_ROWS, IS_KEYS)):
            r = is_row(i, k)
            ws.cell(r, 1, label).font = BOLD
            for yi, y in enumerate(YEARS):
                base = 2 + yi * 5
                cq1, cq2, cq3, cq4, cfy = base, base + 1, base + 2, base + 3, base + 4
                qflow, fy_rec, cumulative = peryear[y]
                # anchor ต่อ "รายการ": ผูก Q4=FY−Q1−Q2−Q3 ได้ก็ต่อเมื่อ FY ของรายการนี้เป็นตัวเลขจริง
                # (กัน circular กับ FY=ผลรวม เมื่อบางรายการเช่นกำไรส่วนแม่ไม่เปิดเผยบางปี)
                fy_val_ok = fy_rec is not None and isinstance(fy_rec.get(key), (int, float))
                std_q4 = fy_val_ok and not cumulative
                for qi, col in enumerate((cq1, cq2, cq3)):
                    rec = qflow.get(("Q1", "Q2", "Q3")[qi])
                    v = rec.get(key) if rec else None
                    ws.cell(r, col).value = round(v, 3) if isinstance(v, (int, float)) else None
                    if v is None:
                        ws.cell(r, col).fill = INPUT_FILL
                if std_q4:
                    ws.cell(r, cq4).value = (f"={GL(cfy)}{r}-{GL(cq1)}{r}"
                                             f"-{GL(cq2)}{r}-{GL(cq3)}{r}")
                    ws.cell(r, cq4).fill = DERIVED_FILL
                else:
                    r4 = qflow.get("Q4")
                    v4 = r4.get(key) if r4 else None
                    ws.cell(r, cq4).value = round(v4, 3) if isinstance(v4, (int, float)) else None
                    if v4 is None:
                        ws.cell(r, cq4).fill = INPUT_FILL
                if fy_val_ok:
                    ws.cell(r, cfy).value = round(fy_rec[key], 3)
                else:                       # FY = ผลรวม 4 ไตรมาส (Q4 เป็นค่า ไม่ใช่สูตร -> ไม่วน)
                    ws.cell(r, cfy).value = (f"={GL(cq1)}{r}+{GL(cq2)}{r}"
                                             f"+{GL(cq3)}{r}+{GL(cq4)}{r}")
                    ws.cell(r, cfy).fill = DERIVED_FILL
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 26
    ws.freeze_panes = "B2"
    return ws


# ---------- Sheet: BS_Timeline (ทุกบริษัท) ----------
def build_bs_timeline(wb, alldata):
    ws = wb.create_sheet("BS_Timeline")
    hdr = ["รายการ (ยอดคงเหลือ, ล้านบาท)", f"31-Dec-{YEARS[0]-1}"]
    for y in YEARS:
        hdr += [f"31-Mar-{y}", f"30-Jun-{y}", f"30-Sep-{y}", f"31-Dec-{y}"]
    ncol = len(hdr)
    ws.append(hdr)
    for i, t in enumerate(TICKERS):
        data = alldata[t]
        title_row(ws, bs_title(i), f"■ {t}", ncol)
        for k, (label, key) in enumerate(zip(BS_ROWS, BS_KEYS)):
            r = bs_row(i, k)
            ws.cell(r, 1, label).font = BOLD
            for y in YEARS:
                recs = data.get(y, {})
                for tag, mi in [("Q1", 0), ("Q2", 1), ("Q3", 2)]:
                    rec = recs.get(tag)
                    if rec and rec.get(key + "_end") is not None:
                        ws.cell(r, BQ.bs_col(y, mi)).value = round(rec[key + "_end"], 3)
                fy = recs.get("FY")
                if fy and fy.get(key + "_end") is not None:
                    ws.cell(r, BQ.bs_col(y, 3)).value = round(fy[key + "_end"], 3)
                src = recs.get("Q1") or recs.get("FY")
                if src and src.get(key + "_begin") is not None:
                    pc = BQ.prev_dec_col(y)
                    if ws.cell(r, pc).value is None:
                        ws.cell(r, pc).value = round(src[key + "_begin"], 3)
            for c in range(2, ncol + 1):
                if ws.cell(r, c).value is None:
                    ws.cell(r, c).fill = INPUT_FILL
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 22
    ws.freeze_panes = "B2"
    return ws


# ---------- Sheet: Quarterly_Ratios (ทุกบริษัท, สูตรสด) ----------
def build_quarterly_ratios(wb):
    ws = wb.create_sheet("Quarterly_Ratios")
    hdr = ["ตัวชี้วัด"] + [f"{q}/{y}" for y in YEARS for q in QTRS]
    ncol = len(hdr)
    ws.append(hdr)
    pct_rows = {3, 4, 9, 10}                  # index ใน RATIO_LABELS ที่เป็น %
    for i, t in enumerate(TICKERS):
        title_row(ws, rt_title(i), f"■ {t}", ncol)
        for k, label in enumerate(RATIO_LABELS):
            ws.cell(rt_row(i, k), 1, label).font = BOLD
        for q in range(20):
            yi, qi = q // 4, q % 4
            col = 2 + q
            Q = GL(col)
            d = DAYS[qi]
            ic = 2 + yi * 5 + qi
            I = GL(ic)
            # อ้างแถวของบริษัท i ในชีต IS_Flow / BS_Timeline
            rev = f"IS_Flow!{I}{is_row(i, 0)}"
            oth = f"IS_Flow!{I}{is_row(i, 1)}"
            cogs = f"IS_Flow!{I}{is_row(i, 2)}"
            np_ = f"IS_Flow!{I}{is_row(i, 3)}"
            bb, be = GL(2 + q), GL(3 + q)

            def avg(kk):
                rr = bs_row(i, kk)
                return f"((BS_Timeline!{bb}{rr}+BS_Timeline!{be}{rr})/2)"

            R = lambda k: f"{Q}{rt_row(i, k)}"     # อ้างเซลล์ในชีตนี้เอง (คอลัมน์เดียวกัน)
            put = lambda k, formula: ws.cell(rt_row(i, k), col, formula)
            put(0, f"={rev}")
            put(1, f"={cogs}")
            put(2, f"={np_}")
            put(3, f'=IFERROR({np_}/{rev},"")')
            put(4, f'=IFERROR(({rev}-{oth}-{cogs})/({rev}-{oth}),"")')
            put(5, f'=IFERROR({avg(0)}/{rev}*{d},"")')
            put(6, f'=IFERROR({avg(1)}/{cogs}*{d},"")')
            put(7, f'=IFERROR({avg(2)}/{cogs}*{d},"")')
            put(8, f'=IFERROR({R(5)}+{R(6)}-{R(7)},"")')
            put(9, f'=IFERROR({np_}/{avg(3)},"")')
            put(10, f'=IFERROR({np_}/{avg(4)},"")')
            put(11, f'=IFERROR(BS_Timeline!{be}{bs_row(i,5)}/BS_Timeline!{be}{bs_row(i,4)},"")')
            for k in range(len(RATIO_LABELS)):
                cell = ws.cell(rt_row(i, k), col)
                cell.fill = FORMULA_FILL
                cell.number_format = "0.0%" if k in pct_rows else "0.0"
    style_header(ws, 1)
    ws.column_dimensions["A"].width = 22
    for q in range(20):
        ws.column_dimensions[GL(2 + q)].width = 10
    ws.freeze_panes = "B2"
    return ws


# ---------- FY section (ใช้ตรรกะเดียวกับ report_linked) ----------
FY_ROWS = RL.INPUT_ROWS
FY_RIDX = {key: i + 3 for i, (key, _) in enumerate(FY_ROWS)}   # +3: title row1, header row2


def build_fy_inputs(wb, groups):
    ws = wb.create_sheet("FY_Inputs")
    ws.append(["ตัวเลขดิบรายปี FY2568 (ล้านบาท) — ช่องเหลืองแก้ได้"])
    ws.append(["รายการ"] + TICKERS)
    for key, label in FY_ROWS:
        row = [label]
        for t in TICKERS:
            iv = RL.input_values(groups[t]) if "FY" in groups[t] else None
            v = iv.get(key) if iv else None
            row.append(round(v, 3) if isinstance(v, (int, float)) else None)
        ws.append(row)
    for c in ws[2]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center")
    ws["A1"].font = BOLD
    for r in range(3, 3 + len(FY_ROWS)):
        ws.cell(r, 1).font = BOLD
        for ci in range(2, 2 + len(TICKERS)):
            ws.cell(r, ci).fill = INPUT_FILL
    ws.column_dimensions["A"].width = 42
    ws.freeze_panes = "B3"
    return ws


def build_summary_fy(wb):
    ws = wb.create_sheet("Summary_FY")
    ws.append(["อัตราส่วนรายปี FY2568 — ฐาน CFA (=สูตรสด อ้าง FY_Inputs)"])
    ws.append(["อัตราส่วน"] + TICKERS)

    def col_of(t):
        return GL(2 + TICKERS.index(t))

    def cell(key, t):
        return f"FY_Inputs!{col_of(t)}{FY_RIDX[key]}"

    def avg(a, b, t):
        return f"(({cell(a, t)}+{cell(b, t)})/2)"

    def op_rev(t):
        return f"({cell('total_revenue', t)}-{cell('other_income', t)})"

    rows = [
        ("NPM (%)", lambda t: f'=IFERROR({cell("net_profit_total", t)}/{cell("total_revenue", t)},"")', True),
        ("Gross Margin (%)", lambda t: f'=IFERROR(({op_rev(t)}-{cell("cogs", t)})/{op_rev(t)},"")', True),
        ("ROA (%)", lambda t: f'=IFERROR({cell("net_profit_total", t)}/{avg("total_assets_begin","total_assets_end", t)},"")', True),
        ("ROE รวม (%)", lambda t: f'=IFERROR({cell("net_profit_total", t)}/{avg("total_equity_begin","total_equity_end", t)},"")', True),
        ("ROE ส่วนแม่ (%)", lambda t: f'=IFERROR({cell("net_profit_parent", t)}/{avg("total_equity_begin","total_equity_end", t)},"")', True),
        ("DSO (วัน)", lambda t: f'=IFERROR({avg("trade_receivables_begin","trade_receivables_end", t)}/{cell("total_revenue", t)}*365,"")', False),
        ("DIO (วัน)", lambda t: f'=IFERROR({avg("inventory_begin","inventory_end", t)}/{cell("cogs", t)}*365,"")', False),
        ("DPO (วัน)", lambda t: f'=IFERROR({avg("trade_payables_begin","trade_payables_end", t)}/{cell("cogs", t)}*365,"")', False),
        ("CCC (วัน)", None, False),
        ("D/E (เท่า)", lambda t: f'=IFERROR({cell("total_liabilities_end", t)}/{cell("total_equity_end", t)},"")', False),
    ]
    row_of = {name: i + 3 for i, (name, _, _) in enumerate(rows)}
    dso_r, dio_r, dpo_r = row_of["DSO (วัน)"], row_of["DIO (วัน)"], row_of["DPO (วัน)"]
    for name, fn, is_pct in rows:
        line = [name]
        for ci, t in enumerate(TICKERS):
            rc = GL(2 + ci)
            if name.startswith("CCC"):
                line.append(f'=IFERROR({rc}{dso_r}+{rc}{dio_r}-{rc}{dpo_r},"")')
            else:
                line.append(fn(t))
        ws.append(line)
    for c in ws[2]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center")
    ws["A1"].font = BOLD
    for ri, (name, _, is_pct) in enumerate(rows):
        r = ri + 3
        ws.cell(r, 1).font = BOLD
        for ci in range(2, 2 + len(TICKERS)):
            cc = ws.cell(r, ci)
            cc.fill = FORMULA_FILL
            cc.number_format = "0.0%" if is_pct else "0.0"
    ws.column_dimensions["A"].width = 20
    for ci in range(2, 2 + len(TICKERS)):
        ws.column_dimensions[GL(ci)].width = 10
    ws.freeze_panes = "B3"
    return ws


def build_notes(wb, alldata):
    ws = wb.create_sheet("Notes")
    have = []
    for t in TICKERS:
        n = sum(1 for y in YEARS for tag in ("Q1", "Q2", "Q3", "FY") if tag in alldata[t].get(y, {}))
        have.append(f"{t}={n}")
    lines = [
        "TOUR — สมุดงานรวมทุกบริษัท (สูตรสด) 5 ปี × 4 ไตรมาส + สรุปรายปี",
        "",
        "ชีตในไฟล์:",
        "  Summary_FY       อัตราส่วนรายปี FY2568 ทุกบริษัท (=สูตร อ้าง FY_Inputs)",
        "  FY_Inputs        ตัวเลขดิบรายปี (ช่องเหลือง แก้ได้)",
        "  Quarterly_Ratios 7 ตัวชี้วัด+GM+D/E × 20 ไตรมาส ทุกบริษัท (=สูตร อ้าง IS_Flow/BS_Timeline)",
        "  IS_Flow          งบกำไรขาดทุน 3 เดือน (Q4 = =FY−Q1−Q2−Q3)",
        "  BS_Timeline      ยอดคงเหลือ 21 จุดเวลา (chain averaging)",
        "",
        "สีของช่อง:  เหลือง = เติมตัวเลขดิบ (แก้ได้)   เขียว = สูตรสด   ส้ม = Q4 (สูตรลบ)",
        "แต่ละบริษัทมีบล็อกของตัวเอง (แถวขึ้นต้น ■ TICKER) เรียงต่อกันในแนวตั้ง",
        "",
        "สูตร (ฐาน CFA; งบดุลเฉลี่ยแบบ chain ปลายไตรมาสก่อน→ปลายไตรมาสนี้; วัน 90/91/92/92; รายปี 365):",
        "  NPM = กำไรสุทธิ / รายได้รวม",
        "  Gross Margin = (รายได้ดำเนินงาน − ต้นทุนขาย) / รายได้ดำเนินงาน   [รายได้ดำเนินงาน = รายได้รวม − รายได้อื่น]",
        "  DSO = เฉลี่ยลูกหนี้การค้า / รายได้รวม × วัน",
        "  DIO = เฉลี่ยสินค้าคงเหลือ / ต้นทุนขาย × วัน",
        "  DPO = เฉลี่ยเจ้าหนี้การค้า / ต้นทุนขาย × วัน",
        "  CCC = DSO + DIO − DPO",
        "  ROA = กำไรสุทธิ / เฉลี่ยสินทรัพย์รวม     ROE = กำไรสุทธิ / เฉลี่ยส่วนของเจ้าของ",
        "  D/E = หนี้สินรวม / ส่วนของเจ้าของรวม (ยอดปลายงวด)",
        "",
        "จำนวนงวดที่มีข้อมูลจริงแล้ว (จากทั้งหมด 20 ไตรมาส): " + ", ".join(have),
        "ช่องเหลืองที่ยังว่าง = ปี 2564–2567 รอเติมงบ; เติมเมื่อไร อัตราส่วนคำนวณเองทันที",
    ]
    for ln in lines:
        ws.append([ln])
    ws.column_dimensions["A"].width = 108
    ws["A1"].font = BOLD
    return ws


def main():
    alldata = {t: BQ.load_company(t) for t in TICKERS}
    groups = RL.load(2568)                    # {ticker: {tag: rec}} ปี 2568 (สำหรับ FY section)
    for t in TICKERS:
        groups.setdefault(t, {})

    wb = Workbook()
    wb.remove(wb.active)
    build_summary_fy(wb)
    build_fy_inputs(wb, groups)
    build_quarterly_ratios(wb)
    build_is_flow(wb, alldata)
    build_bs_timeline(wb, alldata)
    build_notes(wb, alldata)

    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, "TOUR_all_linked.xlsx")
    wb.save(path)
    print("wrote", os.path.relpath(path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
