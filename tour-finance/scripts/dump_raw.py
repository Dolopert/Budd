#!/usr/bin/env python3
"""
dump_raw.py — ส่งออก "ตัวเลขดิบต่อไฟล์งบ" (ก่อน de-cumulate / derive Q4 / คำนวณอัตราส่วน)

ต่างจาก metrics_all.csv (ผ่านการรวมแล้ว): ไฟล์นี้คือค่าที่ extract_set ดึงจากแต่ละไฟล์งบตรงๆ
1 แถว = 1 ไฟล์งบ — เห็นยอดตามที่ปรากฏจริง (งบสะสมก็ยังเป็นยอดสะสม)
หน่วยถูกแปลงเป็นล้านบาทแล้ว

ใช้:  python3 dump_raw.py     ->  output/raw_extract.xlsx
"""
import os
import sys
import glob

sys.path.insert(0, os.path.dirname(__file__))
import extract_set as E
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "output"))
RAW = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data", "raw"))

COLS = [
    ("ticker", "บริษัท"), ("year", "ปี"), ("tag", "งวด(ไฟล์)"),
    ("is_cumulative", "สะสม?"), ("unit", "หน่วยเดิม"),
    ("total_revenue", "รายได้รวม"), ("other_income", "รายได้อื่น"), ("cogs", "ต้นทุนขาย"),
    ("net_profit_total", "กำไรสุทธิรวม"), ("net_profit_parent", "กำไรส่วนแม่"),
    ("trade_receivables_begin", "ลูกหนี้ต้นงวด"), ("trade_receivables_end", "ลูกหนี้ปลายงวด"),
    ("inventory_begin", "สินค้าต้นงวด"), ("inventory_end", "สินค้าปลายงวด"),
    ("trade_payables_begin", "เจ้าหนี้ต้นงวด"), ("trade_payables_end", "เจ้าหนี้ปลายงวด"),
    ("total_assets_end", "สินทรัพย์รวม"), ("total_equity_end", "ส่วนเจ้าของ"),
    ("total_liabilities_end", "หนี้สินรวม"), ("balance_ok", "balance"), ("file", "ไฟล์"),
]
ORDER = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4, "FY": 5}


def main():
    files = []
    for ext in ("xls", "xlsx", "xlsm", "XLS", "XLSX"):
        files += glob.glob(os.path.join(RAW, "*", "*", f"*.{ext}"))
    rows = []
    for f in sorted(set(files)):
        try:
            r = E.extract_file(f)
        except E.ExtractError:
            continue
        r["tag"] = "FY" if r["is_annual"] else r["quarter"]
        r["unit"] = {1e6: "บาท", 1e3: "พันบาท", 1.0: "ล้านบาท"}.get(r["pl_unit_div"], "?")
        rows.append(r)
    rows.sort(key=lambda r: (r["ticker"], r["year"], ORDER.get(r["tag"], 9)))

    wb = Workbook()
    ws = wb.active
    ws.title = "raw_extract_per_file"
    ws.append([c[1] for c in COLS])
    for r in rows:
        ws.append([round(r.get(k), 2) if isinstance(r.get(k), float) else r.get(k, "")
                   for k, _ in COLS])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "D2"
    for i in range(1, len(COLS) + 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 13
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["C"].width = 9
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, "raw_extract.xlsx")
    wb.save(path)
    print(f"wrote {os.path.relpath(path)} — {len(rows)} ไฟล์งบ (ดิบต่อไฟล์ ก่อน de-cumulate/derive Q4)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
