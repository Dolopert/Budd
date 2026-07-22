#!/usr/bin/env python3
"""
build_5yr.py — สร้างโครง Excel 5 ปี × 4 ไตรมาส สำหรับหุ้นกลุ่ม TOUR

ต่างจากต้นแบบ MINT v4 (ที่มี formula เฉพาะปี 2568): เวอร์ชันนี้วาง formula
ครบทั้ง 20 ไตรมาส เชื่อมไปยัง input sheets — เติมข้อมูลปีไหน คอลัมน์นั้นคำนวณเอง

โครง 7 ชีต (ตรงกับ MINT_test_5yr_v4.xlsx):
  1. MDA_Cumulative_Input   — ตัวเลขสะสมจาก MD&A (20 คอลัมน์: 5ปี × Q1/6M/9M/FY)
  2. IncomeStatement_Actual — รายได้ + ต้นทุนขายรายไตรมาสจริง (Q4 = FY - Q1-Q2-Q3)
  3. BalanceSheet_Input     — ยอดคงเหลือ 21 จุดเวลา (ต้นงวด + 20 สิ้นไตรมาส)
  4. <TICKER>_Quarterly     — ผลลัพธ์ 7 ตัวชี้วัด × 20 ไตรมาส (formula)
  5. Methodology_CFA
  6. Methodology_ReverseEng
  7. Notes

ใช้:
    python3 build_5yr.py <TICKER>            # สร้าง output/<TICKER>_5yr.xlsx
    python3 build_5yr.py --all               # สร้างครบทั้ง 10 บริษัท
"""
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TICKERS = ["ASIA", "CENTEL", "DUSIT", "ERW", "MANRIN", "MINT", "OHTL", "SHANG", "SHR", "VRANDA"]
YEARS = [2564, 2565, 2566, 2567, 2568]           # พ.ศ.
QTRS = ["Q1", "Q2", "Q3", "Q4"]
DAYS = [90, 91, 92, 92]                            # จำนวนวันในไตรมาส Q1..Q4

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
LABEL_FONT = Font(bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # เหลือง = ช่องเติมข้อมูล
PLACEHOLDER = "รอไฟล์"

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill, cell.font = HDR_FILL, HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)


# ---------- Sheet 1: MDA_Cumulative_Input ----------
def build_mda(wb):
    ws = wb.create_sheet("MDA_Cumulative_Input")
    hdr = ["รายการ (สะสม ณ สิ้นงวด)"]
    for y in YEARS:
        hdr += [f"{y}_Q1(3m)", f"{y}_6M", f"{y}_9M", f"{y}_FY(12m)"]
    hdr += ["แหล่งอ้างอิง"]
    ws.append(hdr)
    rows = [
        ("รายได้รวม (ตามที่รายงาน, ล้านบาท)", 'MD&A - ตาราง "ผลการดำเนินงาน" (ตามที่รายงาน)'),
        ("อัตรากำไรขั้นต้น (%, สะสม)", 'MD&A - ตาราง "อัตราส่วนทางการเงิน"'),
        ("EBITDA (Core, สะสม, ล้านบาท)", 'MD&A - "ผลการดำเนินงาน" (จากการดำเนินงาน)'),
        ("กำไรสุทธิ (Core, สะสม, ล้านบาท)", 'MD&A - "ผลการดำเนินงาน" (จากการดำเนินงาน)'),
    ]
    for label, ref in rows:
        r = [label] + [PLACEHOLDER] * 20 + [ref]
        ws.append(r)
    style_header(ws)
    _mark_labels_and_inputs(ws, ncols=20)
    return ws


# ---------- Sheet 2: IncomeStatement_Actual ----------
def build_income(wb):
    ws = wb.create_sheet("IncomeStatement_Actual")
    hdr = ["รายการ (ล้านบาท)"]
    for y in YEARS:
        hdr += [f"Q1/{y}", f"Q2/{y}", f"Q3/{y}", f"Q4/{y}(=FY-Q1-Q2-Q3)", f"FY{y}"]
    hdr += ["แหล่งอ้างอิง"]
    ws.append(hdr)
    # 2 rows: revenue, COGS. Q4 = FY - Q1 - Q2 - Q3 (formula)
    for label, ref in [
        ("รวมรายได้ (ตามงบกำไรขาดทุน)", 'งบกำไรขาดทุน - บรรทัด "รวมรายได้"'),
        ("ต้นทุนขาย (ต้นทุนตรงโรงแรม+อาหารฯ)", 'งบกำไรขาดทุน - ต้นทุนโดยตรงโรงแรม + ต้นทุนอาหารและเครื่องดื่ม'),
    ]:
        ws.append([label])
    # fill cells per year block
    for ri in (2, 3):
        for yi in range(5):
            base = 2 + yi * 5           # col of Q1
            c_q1, c_q2, c_q3, c_q4, c_fy = base, base + 1, base + 2, base + 3, base + 4
            for c in (c_q1, c_q2, c_q3, c_fy):
                ws.cell(ri, c).value = PLACEHOLDER
            L = get_column_letter
            ws.cell(ri, c_q4).value = f"={L(c_fy)}{ri}-{L(c_q1)}{ri}-{L(c_q2)}{ri}-{L(c_q3)}{ri}"
        ws.cell(ri, 2 + 5 * 5).value = {2: 'งบกำไรขาดทุน - "รวมรายได้"',
                                        3: "งบกำไรขาดทุน - ต้นทุนตรง+อาหาร"}[ri]
    for r in (2, 3):
        ws.cell(r, 1).font = LABEL_FONT
    style_header(ws)
    return ws


# ---------- Sheet 3: BalanceSheet_Input ----------
def build_balance(wb):
    ws = wb.create_sheet("BalanceSheet_Input")
    # 21 date columns: begin (end of prior year) + 20 quarter-ends
    hdr = ["รายการ (ยอดคงเหลือ ณ วันที่, ล้านบาท)", f"31-Dec-{YEARS[0]-1}(ต้นงวด)"]
    for y in YEARS:
        hdr += [f"31-Mar-{y}", f"30-Jun-{y}", f"30-Sep-{y}", f"31-Dec-{y}"]
    hdr += ["แหล่งอ้างอิง"]
    ws.append(hdr)
    rows = [
        "ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น - สุทธิ",
        "สินค้าคงเหลือ",
        "เจ้าหนี้การค้าและเจ้าหนี้หมุนเวียนอื่น",
        "สินทรัพย์รวม",
        "ส่วนของเจ้าของรวม",
    ]
    for label in rows:
        ws.append([label] + [PLACEHOLDER] * 21 + ["งบฐานะการเงิน (SET-filed CSV) แต่ละไตรมาส"])
    style_header(ws)
    _mark_labels_and_inputs(ws, ncols=21)
    return ws


# ---------- Sheet 4: <TICKER>_Quarterly ----------
def build_quarterly(wb, ticker):
    ws = wb.create_sheet(f"{ticker}_Quarterly")
    hdr = ["รายการ"] + [f"{q}/{y}" for y in YEARS for q in QTRS]
    ws.append(hdr)
    L = get_column_letter

    labels = [
        "รายได้รวม (ล้านบาท)",                         # 2
        "ต้นทุนขาย (จริงจากงบกำไรขาดทุน, ล้านบาท)",     # 3
        "กำไรขั้นต้น (ล้านบาท)",                        # 4
        "อัตรากำไรขั้นต้น (%)",                          # 5
        "EBITDA (Core, ล้านบาท)",                       # 6
        "EBITDA Margin (%)",                            # 7
        "กำไรสุทธิ (Core, ล้านบาท)",                    # 8
        "อัตรากำไรสุทธิ NPM (%)",                        # 9
        "DSO (วัน)",                                    # 10
        "DIO (วัน)",                                    # 11
        "DPO (วัน)",                                    # 12
        "CCC (วัน)",                                    # 13
        "ROA รายไตรมาส (%)",                            # 14
        "ROE รายไตรมาส (%)",                            # 15
    ]
    for lab in labels:
        ws.append([lab])

    for q in range(20):
        yi, qi = q // 4, q % 4
        col = 2 + q                       # column in Quarterly sheet (B=2 .. U=21)
        Q = L(col)
        d = DAYS[qi]

        # -- Income Statement references (5-col blocks per year) --
        inc_base = 2 + yi * 5
        rev_col = L(inc_base + qi)        # Q1..Q4 within year
        cogs_col = rev_col                # same column, row 3 in IncomeStatement
        rev = f"IncomeStatement_Actual!{rev_col}2"
        cogs = f"IncomeStatement_Actual!{cogs_col}3"

        # -- MD&A cumulative references (de-cumulation) --
        mda_base = 2 + yi * 4            # col of Q1(3m) cumulative
        cum = [L(mda_base + k) for k in range(4)]   # Q1,6M,9M,FY
        def decum(row):
            if qi == 0:
                return f"MDA_Cumulative_Input!{cum[0]}{row}"
            return f"MDA_Cumulative_Input!{cum[qi]}{row}-MDA_Cumulative_Input!{cum[qi-1]}{row}"

        # -- Balance sheet references: begin=col(2+q), end=col(3+q) --
        b_beg, b_end = L(2 + q), L(3 + q)
        def avg(row):
            return f"((BalanceSheet_Input!{b_beg}{row}+BalanceSheet_Input!{b_end}{row})/2)"

        ws.cell(2, col).value = f"={rev}"
        ws.cell(3, col).value = f"={cogs}"
        ws.cell(4, col).value = f"={Q}2-{Q}3"
        ws.cell(5, col).value = f'=IFERROR({Q}4/{Q}2,"")'
        ws.cell(6, col).value = f"={decum(4)}"                     # EBITDA (MDA row4)
        ws.cell(7, col).value = f'=IFERROR({Q}6/{Q}2,"")'
        ws.cell(8, col).value = f"={decum(5)}"                     # Net profit (MDA row5)
        ws.cell(9, col).value = f'=IFERROR({Q}8/{Q}2,"")'
        ws.cell(10, col).value = f'=IFERROR({avg(2)}/{Q}2*{d},"")'   # DSO: AR/rev
        ws.cell(11, col).value = f'=IFERROR({avg(3)}/{Q}3*{d},"")'   # DIO: inv/cogs
        ws.cell(12, col).value = f'=IFERROR({avg(4)}/{Q}3*{d},"")'   # DPO: ap/cogs
        ws.cell(13, col).value = f"={Q}11+{Q}10-{Q}12"               # CCC
        ws.cell(14, col).value = f'=IFERROR({Q}8/{avg(5)},"")'       # ROA: np/avg assets
        ws.cell(15, col).value = f'=IFERROR({Q}8/{avg(6)},"")'       # ROE: np/avg equity

    # reverse-eng reference block (per-year self-reported ratios, filled manually)
    ws.cell(17, 1).value = "อ้างอิง: อัตราส่วนรายปีที่บริษัทเปิดเผยตรงใน MD&A (เทียบค่าที่คำนวณเอง)"
    ws.cell(17, 1).font = LABEL_FONT
    for i, name in enumerate(["DSO รายปี (วัน, รายงาน)", "DIO รายปี (วัน, รายงาน)",
                              "DPO รายปี (วัน, รายงาน)", "ROA รายปี (%, รายงาน)",
                              "ROE รายปี (%, รายงาน)"]):
        ws.cell(18 + i, 1).value = name

    for r in range(2, 16):
        ws.cell(r, 1).font = LABEL_FONT
    style_header(ws)
    ws.column_dimensions["A"].width = 34
    return ws


# ---------- Sheets 5-7: static docs ----------
def build_methodology_cfa(wb):
    ws = wb.create_sheet("Methodology_CFA")
    data = [
        ["Methodology 1 — CFA / Damodaran Standard (สูตรมาตรฐานสำหรับงานวิจัย)"],
        ["หลักการ", 'ใช้ "รายได้" เป็นตัวหาร DSO และ "ต้นทุนขาย (COGS)" เป็นตัวหาร DIO/DPO', "CFA Institute; Damodaran (NYU Stern)"],
        ["ค่าเฉลี่ยงบดุล", "ทุกรายการงบดุลใช้ (ยอดต้นงวด + ยอดปลายงวด) / 2", "มาตรฐานสากล"],
        ["จำนวนวัน", "รายไตรมาส Q1=90, Q2=91, Q3=92, Q4=92; รายปี=365", ""],
        ["ตัวชี้วัด", "สูตร", "หมายเหตุ"],
        ["DSO", "ค่าเฉลี่ยลูกหนี้การค้า / รายได้รวม × วัน", 'ใช้ "ลูกหนี้การค้า" ล้วน ถ้ามีจากหมายเหตุ'],
        ["DIO", "ค่าเฉลี่ยสินค้าคงเหลือ / ต้นทุนขาย × วัน", ""],
        ["DPO", "ค่าเฉลี่ยเจ้าหนี้การค้า / ต้นทุนขาย × วัน", 'ใช้ "เจ้าหนี้การค้า" ล้วน ถ้ามีจากหมายเหตุ'],
        ["CCC", "DIO + DSO − DPO", "ค่าติดลบ = ได้เงินก่อนจ่าย (ปกติสำหรับโรงแรม/ร้านอาหาร)"],
        ["ROA", "กำไรสุทธิ / ค่าเฉลี่ยสินทรัพย์รวม", "รายไตรมาส = อัตราต่อไตรมาส (ไม่ annualize)"],
        ["ROE", "กำไรสุทธิ / ค่าเฉลี่ยส่วนของผู้ถือหุ้น", ""],
        ["NPM", "กำไรสุทธิ / รายได้รวม", ""],
        ["ข้อดี", "สูตรมาตรฐานอ้างอิงได้ในวิทยานิพนธ์ + เทียบข้ามบริษัทได้ยุติธรรม (apples-to-apples)", ""],
        ["ข้อจำกัด", "อาจต่างจากที่บริษัทเปิดเผยเอง — ดู Methodology 2 ประกอบ", ""],
    ]
    for row in data:
        ws.append(row)
    ws.cell(1, 1).font = LABEL_FONT
    ws.cell(5, 1).font = LABEL_FONT
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 45
    return ws


def build_methodology_reverse(wb):
    ws = wb.create_sheet("Methodology_ReverseEng")
    data = [
        ["Methodology 2 — Reverse-Engineering Framework (หา assumption ที่บริษัทใช้จริง)"],
        ["วัตถุประสงค์: ทดสอบสูตรหลาย variant เทียบอัตราส่วนรายปีที่บริษัทเปิดเผย แล้วหา variant ที่ error ต่ำสุด"],
        ["เป้าหมาย: หา assumption ร่วมที่ match ดีในหลายบริษัท → ใช้เป็นสูตรกลางของงานวิจัย"],
        ["ตัวชี้วัด", "Variant (สูตรที่ทดสอบ)", "คำนวณ", "เป้า (รายงาน)", "Error", "BEST?"],
        ["DSO", "AR_avg / รายได้รวม × วัน", "", "", "", ""],
        ["DSO", "AR_avg / รายได้ดำเนินงาน × วัน", "", "", "", ""],
        ["DSO", "AR_ปลายงวด / รายได้รวม × วัน", "", "", "", ""],
        ["DIO", "INV_avg / ต้นทุนขาย(แคบ) × วัน", "", "", "", ""],
        ["DIO", "INV_avg / ต้นทุนขาย(รวมโรงแรมอื่น) × วัน", "", "", "", ""],
        ["DIO", "INV_avg / (ต้นทุนขาย+SG&A) × วัน", "", "", "", ""],
        ["DPO", "AP_avg / ต้นทุนขาย(แคบ) × วัน", "", "", "", ""],
        ["DPO", "AP_avg / (ต้นทุนขาย+SG&A) × วัน", "", "", "", ""],
        ["DPO", "AP_ปลายงวด / (ต้นทุนขาย+SG&A) × วัน", "", "", "", ""],
        ["DPO", "AP_avg / ค่าใช้จ่ายรวม(ไม่รวมดอกเบี้ย) × วัน", "", "", "", ""],
        [""],
        ["ขั้นตอนเมื่อได้ครบ 10 บริษัท:"],
        ["1", "รันตารางนี้กับทุกบริษัท เพิ่มคอลัมน์ผลลัพธ์รายบริษัท"],
        ["2", "นับว่า variant ไหนเป็น BEST MATCH ซ้ำบ่อยที่สุดข้ามบริษัท"],
        ["3", "ถ้า variant เดียว match ดีหลายบริษัท → ใช้เป็นสูตรกลาง (assumption ร่วม)"],
        ["4", "ถ้าแต่ละบริษัทคนละ variant → ไม่มี assumption ร่วม กลับไปใช้ Methodology 1 และรายงานส่วนต่าง"],
    ]
    for row in data:
        ws.append(row)
    ws.cell(1, 1).font = LABEL_FONT
    ws.cell(4, 1).font = LABEL_FONT
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    return ws


def build_notes(wb, ticker):
    ws = wb.create_sheet("Notes")
    notes = [
        f"ไฟล์นี้คือ {ticker}_Quarterly แบบ 5 ปี × 4 ไตรมาส (โครง formula ครบ 20 ไตรมาส)",
        "แหล่งข้อมูลต่อไตรมาส: (1) MD&A (2) งบกำไรขาดทุน (3) งบฐานะการเงิน — SET-filed CSV/PDF",
        "EBITDA/กำไรสุทธิ: คำนวณด้วยสูตรลบสะสม (Qn = สะสม_n − สะสม_(n-1)) จาก MD&A",
        "รายได้/ต้นทุนขาย: ใช้ตัวเลขไตรมาสจริงจากงบกำไรขาดทุน; Q4 = FY − Q1 − Q2 − Q3",
        "งบดุล: แปลงหน่วยเป็นล้านบาทให้ตรงกับ MD&A (พันบาท ÷1,000 / บาท ÷1,000,000)",
        "DSO/DIO/DPO: ค่าเฉลี่ยงบดุล ÷ (รายได้ หรือ ต้นทุนขาย) × วันในไตรมาส — สูตรมาตรฐาน CFA",
        "ROA/ROE: กำไรสุทธิ ÷ ค่าเฉลี่ยสินทรัพย์รวม/ส่วนของเจ้าของ (รายไตรมาส ไม่ annualize)",
        "บล็อกท้ายชีต Quarterly (แถว 18+): เติมอัตราส่วนรายปีที่บริษัทเปิดเผยเอง เพื่อ reverse-eng",
        "สถานะ: เติมข้อมูลปีไหนใน input sheets คอลัมน์นั้นใน Quarterly จะคำนวณอัตโนมัติ",
    ]
    for n in notes:
        ws.append([n])
    ws.column_dimensions["A"].width = 100
    return ws


# ---------- helpers ----------
def _mark_labels_and_inputs(ws, ncols):
    """ทำ label หนา + ระบายช่อง input สีเหลือง (คอลัมน์ 2..ncols+1, ทุกแถวข้อมูล)"""
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1)
        if c1.value and not str(c1.value).startswith(("หมายเหตุ", "สีน้ำเงิน")):
            c1.font = LABEL_FONT
        for c in range(2, 2 + ncols):
            cell = ws.cell(r, c)
            if cell.value == PLACEHOLDER:
                cell.fill = INPUT_FILL


def build_company(ticker):
    wb = Workbook()
    wb.remove(wb.active)
    build_mda(wb)
    build_income(wb)
    build_balance(wb)
    build_quarterly(wb, ticker)
    build_methodology_cfa(wb)
    build_methodology_reverse(wb)
    build_notes(wb, ticker)
    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, f"{ticker}_5yr.xlsx")
    wb.save(path)
    return os.path.normpath(path)


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    if argv[1] == "--all":
        for t in TICKERS:
            print("wrote", build_company(t))
        return 0
    t = argv[1].upper()
    if t not in TICKERS:
        print(f"unknown ticker {t}; valid: {', '.join(TICKERS)}")
        return 1
    print("wrote", build_company(t))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
