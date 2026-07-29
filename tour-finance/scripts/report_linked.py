#!/usr/bin/env python3
"""
report_linked.py — เวอร์ชัน "ผูกสูตรสด" ของ ratios_report

ต่างจาก report.py ที่เขียนค่าที่คำนวณเสร็จแล้ว (hardcoded) ลง Excel:
ไฟล์นี้เขียน "ตัวเลขดิบ" (รายได้/ต้นทุนขาย/ลูกหนี้/…) ลงชีต Inputs (ช่องสีเหลือง)
แล้วชีต Ratios คำนวณอัตราส่วนด้วย =สูตร Excel สด ที่อ้างอิงกลับไปยังชีต Inputs

ประโยชน์:
  - ตรวจสอบได้: คลิกที่ช่อง ROE แล้วเห็นสูตร =Inputs!.../((...+...)/2) ตรงๆ
  - แก้ input ช่องเดียว อัตราส่วนที่เกี่ยวข้องอัปเดตเองทันทีใน Excel
  - ใช้เป็น "หลักฐานการคำนวณ" ประกอบวิทยานิพนธ์ (ไม่ใช่กล่องดำ)

ใช้:
    python3 report_linked.py [ปี]        # ค่าเริ่มต้น = 2568
    -> output/ratios_linked_<ปี>.xlsx
"""
import os
import sys
import glob

sys.path.insert(0, os.path.dirname(__file__))
import extract_set as E
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RAW = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data", "raw"))
OUT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "output"))

HDR = PatternFill("solid", fgColor="1F4E78")
HF = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # เหลือง = ช่องตัวเลขดิบ (แก้ได้)
FORMULA_FILL = PatternFill("solid", fgColor="E2EFDA")  # เขียวอ่อน = ช่องสูตร
BOLD = Font(bold=True)

# รายการ input ต่อบริษัท (ล้านบาท) — begin = ยอดต้นงวด (ต้นปี), end = ยอดปลายปี
INPUT_ROWS = [
    ("total_revenue", "รายได้รวม (FY)"),
    ("other_income", "รายได้อื่น (FY)"),
    ("cogs", "ต้นทุนขาย (FY)"),
    ("net_profit_total", "กำไรสุทธิ (รวมส่วนได้เสียไม่มีอำนาจควบคุม)"),
    ("net_profit_parent", "กำไรสุทธิ (ส่วนของผู้ถือหุ้นบริษัทใหญ่)"),
    ("trade_receivables_begin", "ลูกหนี้การค้า — ต้นปี"),
    ("trade_receivables_end", "ลูกหนี้การค้า — ปลายปี"),
    ("inventory_begin", "สินค้าคงเหลือ — ต้นปี"),
    ("inventory_end", "สินค้าคงเหลือ — ปลายปี"),
    ("trade_payables_begin", "เจ้าหนี้การค้า — ต้นปี"),
    ("trade_payables_end", "เจ้าหนี้การค้า — ปลายปี"),
    ("total_assets_begin", "สินทรัพย์รวม — ต้นปี"),
    ("total_assets_end", "สินทรัพย์รวม — ปลายปี"),
    ("total_equity_begin", "ส่วนของเจ้าของรวม — ต้นปี"),
    ("total_equity_end", "ส่วนของเจ้าของรวม — ปลายปี"),
    ("total_liabilities_end", "หนี้สินรวม — ปลายปี"),
]
RIDX = {key: i + 2 for i, (key, _) in enumerate(INPUT_ROWS)}   # แถวใน Inputs (แถว 1 = header)


def load(year):
    """โหลด record รายบริษัทของปีที่ระบุ; คืน {ticker: {tag: rec}}"""
    files = []
    for ext in ("xls", "xlsx", "xlsm", "XLS", "XLSX"):
        files += glob.glob(os.path.join(RAW, "*", "*", f"*.{ext}"))
    groups = {}
    for f in sorted(set(files)):
        try:
            rec = E.extract_file(f)
        except E.ExtractError:
            continue
        if rec["year"] != year:
            continue
        tag = "FY" if rec["is_annual"] else rec["quarter"]
        groups.setdefault(rec["ticker"], {})[tag] = rec
    return groups


def input_values(bt):
    """ดึงค่า input รายบริษัทสำหรับปี: begin = ต้นปี (จากไฟล์ Q1 ถ้ามี), end = ปลายปี (FY)"""
    fy = bt.get("FY")
    if fy is None:
        return None
    beg = bt.get("Q1") or fy         # ยอด 'ต้นปี' อ่านจากคอลัมน์เทียบของไฟล์ Q1
    vals = {}
    for key, _ in INPUT_ROWS:
        if key.endswith("_begin"):
            base = key[:-6]
            vals[key] = beg.get(base + "_begin")
        elif key.endswith("_end"):
            vals[key] = fy.get(key)
        else:
            vals[key] = fy.get(key)
    return vals


def col_letter(i):
    """1-based -> A, B, ... (พอสำหรับ 10 บริษัท)"""
    from openpyxl.utils import get_column_letter
    return get_column_letter(i)


def main(argv):
    year = int(argv[1]) if len(argv) > 1 else 2568
    groups = load(year)
    tickers = sorted(t for t, bt in groups.items() if "FY" in bt)
    if not tickers:
        print(f"ไม่พบข้อมูล FY{year}")
        return 1

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: Inputs (ตัวเลขดิบ, ช่องเหลือง แก้ได้) ----
    wi = wb.create_sheet("Inputs")
    wi.append([f"ตัวเลขดิบจากงบตรวจสอบ FY{year} (ล้านบาท) — ช่องเหลืองแก้ได้"] +
              [""] * len(tickers))
    wi.append(["รายการ"] + tickers)
    for key, label in INPUT_ROWS:
        row = [label]
        for t in tickers:
            v = input_values(groups[t]).get(key)
            row.append(round(v, 3) if isinstance(v, (int, float)) else None)
        wi.append(row)
    for c in wi[2]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center")
    wi["A1"].font = BOLD
    for r in range(3, 3 + len(INPUT_ROWS)):
        wi.cell(r, 1).font = BOLD
        for ci in range(2, 2 + len(tickers)):
            wi.cell(r, ci).fill = INPUT_FILL
    wi.column_dimensions["A"].width = 42
    for ci in range(2, 2 + len(tickers)):
        wi.column_dimensions[col_letter(ci)].width = 11

    # ---- Sheet 2: Ratios (สูตรสด อ้างอิง Inputs) ----
    wr = wb.create_sheet("Ratios")
    wr.append([f"อัตราส่วน FY{year} — ฐาน CFA (=สูตรสด อ้างอิงชีต Inputs)"] + [""] * len(tickers))
    wr.append(["อัตราส่วน"] + tickers)

    def C(ticker):
        """คอลัมน์ของบริษัทในชีต Inputs (ตรงกับตำแหน่งใน tickers)"""
        return col_letter(2 + tickers.index(ticker))

    def cell(key, ticker):
        return f"Inputs!{C(ticker)}{RIDX[key]}"

    # แต่ละอัตราส่วน: (label, ฟังก์ชันสร้างสูตรจาก ticker, เป็น %?)
    def avg(a, b, t):
        return f"(({cell(a, t)}+{cell(b, t)})/2)"

    def op_rev(t):     # รายได้ดำเนินงาน = รายได้รวม − รายได้อื่น
        return f"({cell('total_revenue', t)}-{cell('other_income', t)})"

    metric_rows = [
        ("NPM (%)", lambda t: f"={cell('net_profit_total', t)}/{cell('total_revenue', t)}", True),
        ("Gross Margin (%)",
         lambda t: f"=({op_rev(t)}-{cell('cogs', t)})/{op_rev(t)}", True),
        ("ROA (%)",
         lambda t: f"={cell('net_profit_total', t)}/{avg('total_assets_begin', 'total_assets_end', t)}", True),
        ("ROE รวม (%)",
         lambda t: f"={cell('net_profit_total', t)}/{avg('total_equity_begin', 'total_equity_end', t)}", True),
        ("ROE ส่วนแม่ (%)",
         lambda t: f"={cell('net_profit_parent', t)}/{avg('total_equity_begin', 'total_equity_end', t)}", True),
        ("DSO (วัน)",
         lambda t: f"={avg('trade_receivables_begin', 'trade_receivables_end', t)}/{cell('total_revenue', t)}*365", False),
        ("DIO (วัน)",
         lambda t: f"={avg('inventory_begin', 'inventory_end', t)}/{cell('cogs', t)}*365", False),
        ("DPO (วัน)",
         lambda t: f"={avg('trade_payables_begin', 'trade_payables_end', t)}/{cell('cogs', t)}*365", False),
        ("CCC (วัน) = DSO+DIO−DPO", None, False),   # จะอ้างอิงแถว DSO/DIO/DPO ในชีตนี้เอง
        ("D/E (เท่า)",
         lambda t: f"={cell('total_liabilities_end', t)}/{cell('total_equity_end', t)}", False),
    ]
    # แถวใน Ratios: header=1, header2=2, metric เริ่มแถว 3
    row_of = {name: i + 3 for i, (name, _, _) in enumerate(metric_rows)}
    dso_r = row_of["DSO (วัน)"]
    dio_r = row_of["DIO (วัน)"]
    dpo_r = row_of["DPO (วัน)"]

    for name, fn, is_pct in metric_rows:
        row = [name]
        for ci, t in enumerate(tickers):
            rc = col_letter(2 + ci)
            if name.startswith("CCC"):
                row.append(f"={rc}{dso_r}+{rc}{dio_r}-{rc}{dpo_r}")
            else:
                row.append(fn(t))
        wr.append(row)

    for c in wr[2]:
        if c.value is not None:
            c.fill, c.font = HDR, HF
            c.alignment = Alignment(horizontal="center")
    wr["A1"].font = BOLD
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    for ri, (name, _, is_pct) in enumerate(metric_rows):
        r = ri + 3
        wr.cell(r, 1).font = BOLD
        for ci in range(2, 2 + len(tickers)):
            cc = wr.cell(r, ci)
            cc.fill = FORMULA_FILL
            cc.border = thin
            cc.number_format = "0.0%" if is_pct else "0.0"
    wr.column_dimensions["A"].width = 24
    for ci in range(2, 2 + len(tickers)):
        wr.column_dimensions[col_letter(ci)].width = 11

    # ---- Sheet 3: อธิบายวิธีอ่าน ----
    wn = wb.create_sheet("อ่านอย่างไร")
    for line in [
        "ไฟล์นี้ = เวอร์ชัน 'ผูกสูตรสด' ของ ratios_report",
        "",
        "ชีต Inputs  : ตัวเลขดิบจากงบตรวจสอบ (ช่องเหลือง) — แก้ได้ ถ้าแก้แล้วอัตราส่วนอัปเดตเอง",
        "ชีต Ratios  : ทุกช่องเป็น =สูตร Excel สด ที่อ้างอิงกลับไปชีต Inputs (คลิกดูสูตรได้)",
        "",
        "สูตรที่ผูกไว้ (ฐาน CFA, ค่าเฉลี่ยงบดุลต้นปี-ปลายปี, 365 วัน):",
        "  NPM         = กำไรสุทธิ / รายได้รวม",
        "  Gross Margin= (รายได้ดำเนินงาน − ต้นทุนขาย) / รายได้ดำเนินงาน   [รายได้ดำเนินงาน = รายได้รวม − รายได้อื่น]",
        "  ROA         = กำไรสุทธิ / เฉลี่ย(สินทรัพย์รวม ต้นปี, ปลายปี)",
        "  ROE         = กำไรสุทธิ / เฉลี่ย(ส่วนของเจ้าของ ต้นปี, ปลายปี)",
        "  DSO         = เฉลี่ย(ลูกหนี้การค้า) / รายได้รวม × 365",
        "  DIO         = เฉลี่ย(สินค้าคงเหลือ) / ต้นทุนขาย × 365",
        "  DPO         = เฉลี่ย(เจ้าหนี้การค้า) / ต้นทุนขาย × 365",
        "  CCC         = DSO + DIO − DPO",
        "  D/E         = หนี้สินรวม(ปลายปี) / ส่วนของเจ้าของ(ปลายปี)",
        "",
        "หมายเหตุ: ค่าที่ได้จะตรงกับ Summary_FY ในไฟล์ ratios_report.xlsx ทุกตัว",
        "         (ไฟล์นั้นเป็นค่า hardcoded, ไฟล์นี้เป็นสูตรสดเพื่อตรวจสอบ/แก้ input ได้)",
    ]:
        wn.append([line])
    wn.column_dimensions["A"].width = 100
    wn["A1"].font = BOLD

    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, f"ratios_linked_{year}.xlsx")
    wb.save(path)
    print("wrote", os.path.relpath(path), "—", len(tickers), "บริษัท:", ", ".join(tickers))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
