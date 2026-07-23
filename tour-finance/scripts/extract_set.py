#!/usr/bin/env python3
"""
extract_set.py — ดึงข้อมูลงบการเงินจากไฟล์ SET (.xls) แบบ config-driven

รับไฟล์ .xls ที่ตลาดหลักทรัพย์เผยแพร่ (ชีต BS / PL / CF) แล้ว:
  1. ตรวจหน่วยจาก header อัตโนมัติ (บาท / พันบาท / ล้านบาท) -> normalize เป็นล้านบาท
  2. ตรวจงวด (ไตรมาส/รายปี) + ปี จากหัวข้องบ
  3. เลือกคอลัมน์ "งบการเงินรวม + งวดปัจจุบัน" อัตโนมัติ
  4. map ชื่อรายการผ่าน mapping/line_items.yml (fail-loud ถ้าหา required ไม่เจอ)
  5. ตรวจ balance check: รวมสินทรัพย์ = รวมหนี้สิน + รวมส่วนของผู้ถือหุ้น

ใช้:
    python3 extract_set.py <ไฟล์.xls> [<ไฟล์.xls> ...]      # ดึง + คำนวณ 7 ตัวชี้วัด
    python3 extract_set.py --dir <โฟลเดอร์บริษัท/ปี>          # ดึงทุกไฟล์ในโฟลเดอร์

7 ตัวชี้วัด (คำนวณจาก PL + BS ล้วน ไม่ต้องใช้ MD&A):
    NPM, ROA, ROE, DSO, DIO, DPO, CCC
"""
import os
import re
import io
import sys
import glob
import xlrd
import openpyxl
import yaml

MAP_PATH = os.path.join(os.path.dirname(__file__), "..", "mapping", "line_items.yml")
DAYS = {"Q1": 90, "Q2": 91, "Q3": 92, "Q4": 92}
THAI_MONTH_Q = {"มีนาคม": "Q1", "มิถุนายน": "Q2", "กันยายน": "Q3", "ธันวาคม": "Q4"}


class ExtractError(Exception):
    pass


# ชื่อบริษัท (คีย์เวิร์ด) -> ticker กลุ่ม TOUR
TICKER_KEYWORDS = {
    "ASIA": ["เอเชียโฮเต็ล", "เอเชีย โฮเต็ล", "asia hotel"],
    "CENTEL": ["เซ็นทรัลพลาซา", "central plaza", "centara"],
    "DUSIT": ["ดุสิตธานี", "ดุสิต", "dusit"],
    "ERW": ["เอราวัณ", "erawan"],
    "MANRIN": ["แมนดาริน", "mandarin"],
    "MINT": ["ไมเนอร์ อินเตอร์เนชั่นแนล", "ไมเนอร์", "minor"],
    "OHTL": ["โอเรียนเต็ล", "โอเอชทีแอล", "ohtl", "oriental"],
    "SHANG": ["แชงกรี", "shangri"],
    "SHR": ["เอส โฮเทล", "เอสโฮเทล", "s hotels", "s hotel"],
    "VRANDA": ["วีรันดา", "veranda"],
}


def detect_company(pl_sheet):
    """เดา ticker จากชื่อบริษัทในหัวงบ (5 แถวแรก); None ถ้าไม่รู้จัก"""
    head = " ".join(norm(pl_sheet.cell_value(r, c)).lower()
                    for r in range(min(pl_sheet.nrows, 5))
                    for c in range(min(pl_sheet.ncols, 6)))
    for ticker, kws in TICKER_KEYWORDS.items():
        if any(k.lower() in head for k in kws):
            return ticker
    return None


def load_map():
    with open(MAP_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


def norm(s):
    """normalize label: strip + ยุบช่องว่างซ้ำ"""
    return re.sub(r"\s+", " ", str(s).strip())


def nospace(s):
    """ตัดช่องว่างทั้งหมด สำหรับเทียบ label (บาง label ใส่เว้นวรรคในวงเล็บ)"""
    return re.sub(r"\s+", "", str(s).strip())


# ---------- unified reader: รองรับทั้ง .xls (xlrd) และ .xlsx (openpyxl) ----------
class Grid:
    """หน้าตาชีตแบบ 0-based เหมือนกันไม่ว่าไฟล์ต้นทางเป็น .xls หรือ .xlsx"""
    def __init__(self, name, nrows, ncols, getter):
        self.name = name
        self.nrows = nrows
        self.ncols = ncols
        self._get = getter

    def cell_value(self, r, c):
        v = self._get(r, c)
        return "" if v is None else v


def _is_xlsx(path):
    """ดูจาก magic bytes ไม่ใช่นามสกุล (บางไฟล์ .XLS จริงๆ เป็น xlsx/zip)"""
    with open(path, "rb") as f:
        sig = f.read(4)
    if sig[:2] == b"PK":                 # zip -> xlsx/xlsm
        return True
    if sig == b"\xD0\xCF\x11\xE0":       # OLE2 -> xls เก่า
        return False
    return path.lower().endswith((".xlsx", ".xlsm"))


def open_book(path):
    """คืน list ของ Grid (0-based) จากไฟล์ .xls หรือ .xlsx (ตรวจชนิดจาก magic bytes)"""
    if _is_xlsx(path):
        # เปิดผ่าน BytesIO เพื่อเลี่ยงที่ openpyxl ปฏิเสธไฟล์ xlsx ที่ตั้งนามสกุล .xls
        with open(path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
        grids = []
        for ws in wb.worksheets:
            nr, nc = ws.max_row or 0, ws.max_column or 0
            data = [[cell.value for cell in row] for row in ws.iter_rows()]

            def make_getter(d):
                def g(r, c):
                    if 0 <= r < len(d) and 0 <= c < len(d[r]):
                        return d[r][c]
                    return None
                return g
            grids.append(Grid(ws.title, nr, nc, make_getter(data)))
        return grids
    # .xls
    wb = xlrd.open_workbook(path)
    return [Grid(sh.name, sh.nrows, sh.ncols,
                 (lambda s: (lambda r, c: s.cell_value(r, c)))(sh))
            for sh in wb.sheets()]


# ---------- label reading (label อาจอยู่คนละคอลัมน์ตาม indent เช่น ASIA) ----------
def row_label(sheet, r, upto):
    """label ของแถว = ข้อความช่องแรกที่ไม่ว่าง ในคอลัมน์ 0..upto-1 (ซ้ายของคอลัมน์ตัวเลข)"""
    for c in range(max(0, upto)):
        t = norm(sheet.cell_value(r, c))
        if t:
            return t
    return ""


def sheet_has_label(sheet, targets, scan_cols=10):
    """จริงถ้ามีแถวใดที่ label (นับช่องซ้ายสุดที่ไม่ว่าง) ตรงกับ targets"""
    tset = {nospace(t) for t in targets}
    for r in range(sheet.nrows):
        if nospace(row_label(sheet, r, scan_cols)) in tset:
            return True
    return False


# ---------- content-based sheet lookup (ทนต่อชื่อชีตแปลกๆ เช่น 'FS.T') ----------
BS_TOTAL_LABELS = ["รวมสินทรัพย์", "รวมหนี้สิน", "รวมส่วนของผู้ถือหุ้น", "รวมส่วนของเจ้าของ"]


def get_bs_sheets(grids):
    """คืน list ชีตงบดุล — บางบริษัทแยกงบดุลเป็นหลายชีต (DUSIT: สินทรัพย์ / หนี้สิน+ทุน)"""
    sheets = [g for g in grids if sheet_has_label(g, BS_TOTAL_LABELS)]
    if not sheets:                                     # เผื่อชื่อชีต (รวมชื่อไทย)
        sheets = [g for g in grids
                  if norm(g.name).upper().startswith(("BS", "FS"))
                  or "ฐานะการเงิน" in norm(g.name) or "งบดุล" in norm(g.name)]
    if not sheets:
        raise ExtractError(f"ไม่พบชีตงบดุล — มีชีต: {[g.name for g in grids]}")
    return sheets


def get_cf_sheet(grids):
    """ชีตงบกระแสเงินสด — จับด้วยชื่อชีต/บรรทัดเฉพาะ CF (ไม่ใช่บรรทัดค่าเสื่อมที่ก็มีในงบ PL)"""
    for g in grids:                              # 1) ชื่อชีต
        n = norm(g.name)
        if "กระแสเงินสด" in n or "cashflow" in n.lower() or n.upper().startswith(("CF", "CASH")):
            return g
    for g in grids:                              # 2) บรรทัดที่มีเฉพาะในงบกระแสเงินสด
        if sheet_has_label(g, ["กระแสเงินสดจากกิจกรรมดำเนินงาน",
                               "เงินสดสุทธิได้มาจากกิจกรรมดำเนินงาน",
                               "เงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมดำเนินงาน",
                               "กระแสเงินสดจากกิจกรรมด าเนินงาน"]):
            return g
    return None


def find_sum_excluding(sheet, col, aliases, divisor, exclude, row_end=None):
    """ผลรวมบรรทัดที่ label ขึ้นต้นด้วย aliases แต่ไม่มีคำใน exclude (ใช้กับค่าเสื่อม CF)"""
    al = [nospace(a) for a in aliases]
    ex = [nospace(e) for e in exclude]
    limit = row_end if row_end is not None else sheet.nrows
    total = 0.0
    hit = False
    for r in range(limit):
        lab = nospace(row_label(sheet, r, col))
        if not lab or any(e in lab for e in ex):
            continue
        if any(lab.startswith(a) for a in al):
            v = sheet.cell_value(r, col)
            if isinstance(v, (int, float)) and v != 0:
                total += v / divisor
                hit = True
    return total if hit else None


def _is_cumulative_pl(name):
    """ชื่อชีตที่บ่งบอกว่าเป็นงบสะสม (6/9/12 เดือน) — ควรหลีกเลี่ยงถ้ามีงบ 3 เดือน"""
    n = norm(name).lower()
    return any(k in n for k in ["(6", "(9", "6m", "9m", "6month", "9month",
                                "6-month", "9-month", "หกเดือน", "เก้าเดือน", "(6mths", "(9mths"])


def _is_three_month(name):
    n = norm(name).lower()
    return any(k in n for k in ["(3)", "(3m", "3-month", "3month", "3 month", "สามเดือน", "(3 "])


def get_pl_sheet(grids):
    """งบกำไรขาดทุน = ชีตที่มี 'รวมรายได้' หรือชื่อชีตบ่งชี้; ถ้ามีทั้ง 3M/6M เลือก 3 เดือน"""
    pls = [g for g in grids if sheet_has_label(g, ["รวมรายได้", "รายได้รวม"])]
    if not pls:                                        # เผื่อไม่มีบรรทัด 'รวมรายได้' (OHTL) / ชื่อไทย (MANRIN)
        pls = [g for g in grids if norm(g.name).upper().startswith("PL")
               or "กำไรขาดทุน" in norm(g.name)]
    if not pls:
        raise ExtractError(f"ไม่พบชีตงบกำไรขาดทุน — มีชีต: {[g.name for g in grids]}")
    for g in pls:                                      # 1) ชอบชีต 3 เดือนที่ระบุชัด
        if _is_three_month(g.name):
            return g
    non_cum = [g for g in pls if not _is_cumulative_pl(g.name)]   # 2) เลี่ยงงบสะสม
    return (non_cum or pls)[0]


# ---------- unit detection ----------
def detect_unit_divisor(sheet):
    """คืนตัวหารเพื่อแปลงค่าในงบ -> ล้านบาท"""
    for r in range(min(sheet.nrows, 10)):
        for c in range(sheet.ncols):
            t = norm(sheet.cell_value(r, c))
            if "หน่วย" in t or "บาท" in t:
                if "พันบาท" in t:
                    return 1_000.0
                if "ล้านบาท" in t:
                    return 1.0
                if "บาท" in t:
                    return 1_000_000.0
    raise ExtractError("ตรวจหน่วยไม่ได้ (ไม่พบ 'หน่วย ... บาท' ใน header)")


# ---------- header-row detection (ใช้ร่วมกันทั้ง period + column) ----------
DATE_RE = re.compile(r"\d{1,2}\s+\S+\s+(25\d\d)")


def year_of(v):
    """คืนปี พ.ศ. ถ้า cell เป็น 'ตัวบอกงวด' (เลขปีล้วน / 'พ.ศ. 2568' / วันที่เต็ม) มิฉะนั้น None"""
    if isinstance(v, (int, float)) and 2500 <= v <= 2600:
        return int(v)
    t = norm(v)
    m = re.fullmatch(r"(?:พ\.?\s*ศ\.?\s*)?(25\d\d)(?:\.0)?", t)   # '2568' หรือ 'พ.ศ. 2568' (MINT)
    if m:
        return int(m.group(1))
    m = DATE_RE.search(t)                          # '31 มีนาคม 2568'
    return int(m.group(1)) if m else None


def period_header(sheet):
    """หาแถว header ของงวด = แถวแรกที่มี 'ตัวบอกงวด' >= 2 ช่อง
    คืน (row, [cols เรียงซ้าย->ขวา], [ปีของแต่ละ col])
    คอลัมน์ซ้ายสุด = งบรวมงวดปัจจุบัน; ถัดไป = งบรวมงวดเทียบ (ต้นงวด/ปีก่อน)
    """
    for r in range(min(sheet.nrows, 16)):
        cols, yrs = [], []
        for c in range(sheet.ncols):
            y = year_of(sheet.cell_value(r, c))
            if y is not None:
                cols.append(c)
                yrs.append(y)
        if len(cols) >= 2:
            return r, cols, yrs
    raise ExtractError("หาแถว header ของงวดไม่ได้ (ไม่พบเลขปี/วันที่ >=2 ช่อง)")


def detect_cumulative(pl_sheet):
    """จริงถ้างบไตรมาสรายงานแบบสะสม (YTD 6/9 เดือน) — บางบริษัท (MANRIN) ไม่มีคอลัมน์ 3 เดือน"""
    blob = " ".join(norm(pl_sheet.cell_value(r, c))
                    for r in range(min(pl_sheet.nrows, 8)) for c in range(pl_sheet.ncols))
    if any(k in blob for k in ["สำหรับปี", "รอบปี", "สิบสองเดือน"]):
        return False                                   # งบรายปี = FY (จัดการแยก)
    return ("หกเดือน" in blob or "เก้าเดือน" in blob or
            "6 เดือน" in blob or "9 เดือน" in blob)


def detect_period(pl_sheet):
    """คืน (quarter_label, year_be, is_annual) — อ่านปีจากแถว header เท่านั้น
    (กันเลขอื่นที่ไม่ใช่ปีมารบกวน) และดูชนิดงวดจากข้อความหัวตาราง
    """
    _, _, yrs = period_header(pl_sheet)
    year = max(yrs)                                # งวดปัจจุบัน = ปีมากสุดในแถว header
    blob = " ".join(norm(pl_sheet.cell_value(r, c))
                    for r in range(min(pl_sheet.nrows, 16)) for c in range(pl_sheet.ncols))
    is_annual = any(k in blob for k in ["สำหรับปี", "รอบปี", "สิบสองเดือน", "สำหรับงวดสิบสอง"])
    if is_annual:
        return "Q4", year, True
    month = next((m for m in THAI_MONTH_Q if m in blob), None)
    if month is None:
        raise ExtractError("ตรวจเดือนสิ้นงวดไม่ได้จากหัวข้องบ")
    return THAI_MONTH_Q[month], year, False


# ---------- column selection (งบการเงินรวม = คอลัมน์ซ้ายสุดของแต่ละงวด) ----------
def pl_current_consol_col(pl_sheet, year_be):
    """คอลัมน์งบรวม-งวดปัจจุบัน = คอลัมน์ตัวบอกงวดซ้ายสุดที่ปี == year_be"""
    _, cols, yrs = period_header(pl_sheet)
    for c, y in zip(cols, yrs):
        if y == year_be:
            return c
    return cols[0]


def bs_date_cols(bs_sheet, year_be):
    """คืน (col_current, col_begin) ของงบการเงินรวม = สองคอลัมน์งวดซ้ายสุด"""
    _, cols, _ = period_header(bs_sheet)
    return cols[0], cols[1]


# ---------- line-item lookup ----------
def pl_first_statement_end(pl_sheet, label_upto):
    """งบไตรมาสมีงบกำไรขาดทุน 2 ชุด (3 เดือน + สะสม) ในชีตเดียว — คืนแถวสิ้นสุดของชุดแรก
    ใช้ 'รวมรายได้' ครั้งที่ 2; ถ้าไม่มี (MANRIN) ใช้หัวข้องวด 'สำหรับงวด/รอบ' ครั้งที่ 2
    """
    rev_rows = [r for r in range(pl_sheet.nrows)
                if nospace(row_label(pl_sheet, r, label_upto)) in ("รวมรายได้", "รายได้รวม")]
    if len(rev_rows) >= 2:
        return rev_rows[1]
    # ไม่มี 'รวมรายได้' 2 ชุด -> หาหัวข้อ 'งบกำไรขาดทุน' ครั้งที่ 2 (จุดเริ่มงบชุดถัดไป)
    # ไม่นับ '(ต่อ)' ซึ่งเป็นส่วนต่อของงบชุดเดียวกัน (OCI)
    title_rows = []
    for r in range(pl_sheet.nrows):
        blob = " ".join(norm(pl_sheet.cell_value(r, c)) for c in range(min(pl_sheet.ncols, 4)))
        if "งบกำไรขาดทุน" in blob and "ต่อ" not in blob:
            title_rows.append(r)
    return title_rows[1] if len(title_rows) >= 2 else pl_sheet.nrows


def find_value(sheet, col, spec, divisor, row_end=None):
    """คืนค่า (ล้านบาท) ของ canonical item ตาม spec; sum ถ้ากำหนด
    row_end จำกัดขอบเขตแถว (ใช้กันไม่ให้ข้ามไปงบชุดสะสม)
    label อ่านจากช่องซ้ายสุดที่ไม่ว่าง (คอลัมน์ 0..col-1) รองรับ indent หลายคอลัมน์
    """
    match_mode = spec.get("match", "prefix")
    aliases = [nospace(a) for a in spec["aliases"]]
    do_sum = spec.get("sum", False)
    reduce = spec.get("reduce")                    # 'max' = เอาค่ามากสุด (ไม่ใช่บรรทัดแรก)
    limit = row_end if row_end is not None else sheet.nrows
    found = []
    for r in range(limit):
        label = nospace(row_label(sheet, r, col))
        if not label:
            continue
        hit = False
        for an in aliases:
            if match_mode == "exact" and label == an:
                hit = True
            elif match_mode == "prefix" and label.startswith(an):
                hit = True
            if hit:
                break
        if not hit:
            continue
        v = sheet.cell_value(r, col)
        if isinstance(v, (int, float)) and v != 0:
            found.append(v / divisor)
            if not do_sum and not reduce:
                break
    if not found:
        return None
    if do_sum:
        return sum(found)
    if reduce == "max":
        return max(found, key=abs)
    return found[0]


def extract_file(path):
    """ดึง canonical values จากไฟล์ SET หนึ่งไฟล์ -> dict"""
    cfg = load_map()
    grids = open_book(path)
    pl = get_pl_sheet(grids)
    bs_sheets = get_bs_sheets(grids)                # อาจมีหลายชีต (งบดุลแยก)

    q, year, is_annual = detect_period(pl)
    pl_div = detect_unit_divisor(pl)
    pl_col = pl_current_consol_col(pl, year)

    # เตรียม (sheet, unit, col_cur, col_begin) ของงบดุลแต่ละใบ
    bs_ctx = []
    for bs in bs_sheets:
        try:
            cur, beg = bs_date_cols(bs, year)
        except ExtractError:
            continue
        bs_ctx.append((bs, detect_unit_divisor(bs), cur, beg))
    if not bs_ctx:
        raise ExtractError(f"{os.path.basename(path)}: หาคอลัมน์งวดของงบดุลไม่ได้")

    out = {"file": os.path.basename(path), "ticker": detect_company(pl),
           "quarter": q, "year": year, "is_annual": is_annual,
           "is_cumulative": (detect_cumulative(pl) if not is_annual else False),
           "pl_unit_div": pl_div}
    missing = []

    # income statement (เฉพาะงบชุดแรก = 3 เดือน; กันข้ามไปงบสะสม)
    pl_end = pl_first_statement_end(pl, pl_col)
    # รายได้หลัก (revenue_main) ต้องอยู่เหนือบรรทัด 'รวมรายได้' แรก — กันไปโดนรายได้งบสะสม
    rev1 = next((r for r in range(pl_end)
                 if nospace(row_label(pl, r, pl_col)) in ("รวมรายได้", "รายได้รวม")), pl_end)
    for name, spec in cfg["income_statement"].items():
        end = rev1 if name == "revenue_main" else pl_end
        v = find_value(pl, pl_col, spec, pl_div, row_end=end)
        if v is None and spec.get("required"):
            missing.append(f"PL:{name}")
        out[name] = v

    # total revenue: งบ multi-step (DUSIT) 'รวมรายได้' = ผลรวมรายได้อื่นเท่านั้น
    # ถ้ารายได้หลักบรรทัดเดียว > 'รวมรายได้' -> total = รายได้หลัก + รวมรายได้(อื่น)
    # ถ้าไม่มี 'รวมรายได้' เลย (OHTL) -> รายได้หลัก + รายได้อื่น
    R, M = out.get("total_revenue"), out.get("revenue_main")
    oth0 = out.get("other_income") or 0
    if R is not None and M is not None and M > R:
        out["total_revenue"] = M + R
    elif R is None and M is not None:
        out["total_revenue"] = M + oth0
    if out.get("total_revenue") is not None and "PL:total_revenue" in missing:
        missing.remove("PL:total_revenue")             # fallback สำเร็จ

    # COGS: ใช้ subtotal ถ้ามี (ASIA) มิฉะนั้นผลรวม components (CENTEL/ERW); เป็นค่าบวกเสมอ
    cogs = out.get("cogs_subtotal") or out.get("cogs_components")
    if cogs is None:
        missing.append("PL:cogs (ทั้ง subtotal และ components)")
    out["cogs"] = abs(cogs) if cogs is not None else None

    # balance sheet: ค้นข้ามทุกชีตงบดุล (รองรับงบดุลแยกชีต) เก็บยอดปลายงวด+ต้นงวด
    for name, spec in cfg["balance_sheet"].items():
        v_cur = v_beg = None
        for bs, div, cur, beg in bs_ctx:
            v = find_value(bs, cur, spec, div)
            if v is not None:
                v_cur = v
                v_beg = find_value(bs, beg, spec, div)
                break
        if v_cur is None and spec.get("required"):
            missing.append(f"BS:{name}")
        out[name + "_end"] = v_cur
        out[name + "_begin"] = v_beg

    # ค่าเสื่อมราคา จากงบกระแสเงินสด (สำหรับสูตร SET: ต้นทุน = ต้นทุน + ค่าเสื่อม)
    out["depreciation"] = None
    cf = get_cf_sheet(grids)
    if cf is not None:
        try:
            cf_col = pl_current_consol_col(cf, year)
            cf_div = detect_unit_divisor(cf)
            dspec = cfg["cash_flow"]["depreciation"]
            out["depreciation"] = find_sum_excluding(
                cf, cf_col, dspec["aliases"], cf_div, dspec.get("exclude", []))
        except ExtractError:
            pass

    if missing:
        raise ExtractError(f"{os.path.basename(path)}: หา required item ไม่เจอ -> {missing}")

    # balance check: รวมสินทรัพย์ = รวมหนี้สิน + รวมส่วนของผู้ถือหุ้น (ยอดปลายงวด)
    ta, tl, te = out["total_assets_end"], out["total_liabilities_end"], out["total_equity_end"]
    if ta and tl and te:
        diff = abs(ta - (tl + te))
        tol = max(1.0, ta * 0.005)               # ผ่อนปรน 0.5% กันปัดเศษ
        out["balance_ok"] = diff <= tol
        out["balance_diff"] = diff
        if not out["balance_ok"]:
            raise ExtractError(
                f"{os.path.basename(path)}: BALANCE CHECK ล้มเหลว "
                f"(สินทรัพย์ {ta:.1f} != หนี้สิน {tl:.1f} + ทุน {te:.1f}, ต่าง {diff:.1f} ล้านบาท)")
    return out


def days_for(q):
    return DAYS[q]


def metrics(np_, rev, cogs, ar_avg, inv_avg, ap_avg, ta_avg, te_avg, days):
    def safe(a, b):
        return a / b if b else None
    dso = safe(ar_avg, rev) and ar_avg / rev * days
    dio = safe(inv_avg, cogs) and inv_avg / cogs * days
    dpo = safe(ap_avg, cogs) and ap_avg / cogs * days
    ccc = (dio + dso - dpo) if None not in (dio, dso, dpo) else None
    return {
        "NPM": safe(np_, rev), "ROA": safe(np_, ta_avg), "ROE": safe(np_, te_avg),
        "DSO": dso, "DIO": dio, "DPO": dpo, "CCC": ccc,
    }


def compute_quarter(rec):
    """คำนวณ 7 ตัวชี้วัดจาก record (ต้นทุน/รายได้ = งวดปัจจุบัน; งบดุลใช้ค่าเฉลี่ยต้น-ปลายงวด)"""
    def avg(a, b):
        return (a + b) / 2 if (a is not None and b is not None) else (a if b is None else b)
    m = metrics(
        rec["net_profit_total"], rec["total_revenue"], rec["cogs"],
        avg(rec["trade_receivables_begin"], rec["trade_receivables_end"]),
        avg(rec["inventory_begin"], rec["inventory_end"]),
        avg(rec["trade_payables_begin"], rec["trade_payables_end"]),
        avg(rec["total_assets_begin"], rec["total_assets_end"]),
        avg(rec["total_equity_begin"], rec["total_equity_end"]),
        days_for(rec["quarter"]),
    )
    return m


BS_ITEMS = ["trade_receivables", "inventory", "trade_payables", "total_assets", "total_equity"]
FLOW_ITEMS = ["total_revenue", "cogs", "net_profit_total", "net_profit_parent",
              "other_income", "revenue_main"]


def _cumulative_year(by_tag):
    """จริงถ้าปีนี้ยื่นงบแบบสะสม (YTD) — Q2=6เดือน / Q3=9เดือน (MANRIN 2564-66)
    ต่างจากงบไตรมาส SET มาตรฐานที่แต่ละไตรมาสเป็นตัวเลข 3 เดือนแยก
    """
    return any(by_tag.get(t, {}).get("is_cumulative") for t in ("Q2", "Q3"))


def _decumulate(by_tag):
    """แปลงงบสะสม -> ตัวเลข 3 เดือนจริง (ลบยอดสะสมงวดก่อน) เฉพาะรายการ flow
    ยอดงบดุล (stock) เป็น point-in-time อยู่แล้ว ไม่แตะ
    จุดที่ 4 = ไฟล์ Q4 (12 เดือน) หรือ FY ถ้าไม่มี Q4
    """
    seq = []
    for t in ("Q1", "Q2", "Q3"):
        if t in by_tag:
            seq.append(by_tag[t])
    last = by_tag.get("Q4") or by_tag.get("FY")
    if last is not None:
        seq.append(last)
    out = []
    prev = {k: 0.0 for k in FLOW_ITEMS}
    for i, rec in enumerate(seq):
        nr = dict(rec)
        nr["quarter"] = ("Q1", "Q2", "Q3", "Q4")[i]      # จุดที่ 4 = Q4 เสมอ
        nr["is_annual"] = False
        for k in FLOW_ITEMS:
            v = rec.get(k)
            if v is not None:
                nr[k] = v - prev[k]
                prev[k] = v
        out.append(nr)
    return out


def compute_group(by_tag):
    """คำนวณ 7 ตัวชี้วัดของทั้งปี โดย chain ยอดปลายไตรมาสก่อนหน้าเป็นยอดต้นงวด

    งบไตรมาส SET เทียบกับ 'ต้นปี' เสมอ ไม่ใช่ไตรมาสก่อน — จึงต้องต่อลำดับเอง:
      Q1 avg = (ต้นปี, มี.ค.)   Q2 avg = (มี.ค., มิ.ย.)
      Q3 avg = (มิ.ย., ก.ย.)   Q4 avg = (ก.ย., ธ.ค.)
    คืน list ของ (rec, metrics) เรียงตามไตรมาสที่มี
    """
    if _cumulative_year(by_tag):
        seq = _decumulate(by_tag)         # งบสะสม (MANRIN 2564-66) -> 3 เดือนจริง
    else:
        seq = [by_tag[t] for t in ("Q1", "Q2", "Q3") if t in by_tag]
        if "Q4" in by_tag:
            seq.append(by_tag["Q4"])      # Q4 จริง (งบ 3 เดือน) — ใช้ตรงถ้ามี
        else:
            q4 = derive_q4(by_tag)        # ไม่มี Q4 จริง -> derive จาก FY = FY-Q1-Q2-Q3
            if q4:
                seq.append(q4)
    if not seq:
        return []
    # ยอดต้นงวดของไตรมาสแรกในลำดับ = ยอด 'ต้นปี' (คอลัมน์เทียบของไฟล์ Q1)
    prev = {it: seq[0].get(it + "_begin") for it in BS_ITEMS}
    out = []
    for rec in seq:
        cur = {it: rec.get(it + "_end") for it in BS_ITEMS}
        avg = {it: _avg(prev[it], cur[it]) for it in BS_ITEMS}
        m = metrics(
            rec["net_profit_total"], rec["total_revenue"], rec["cogs"],
            avg["trade_receivables"], avg["inventory"], avg["trade_payables"],
            avg["total_assets"], avg["total_equity"], days_for(rec["quarter"]),
        )
        out.append((rec, m))
        prev = cur
    return out


def _avg(a, b):
    if a is None and b is None:
        return None
    if a is None:
        return b
    if b is None:
        return a
    return (a + b) / 2


def fmt(v, pct=False):
    if v is None:
        return "—"
    return f"{v*100:.1f}%" if pct else f"{v:.1f}"


def derive_q4(recs_by_tag):
    """สร้าง record ไตรมาส 4 = FY − Q1 − Q2 − Q3 (งบดุล Q4 ใช้ปลายปี = FY)"""
    need = ["Q1", "Q2", "Q3", "FY"]
    if any(t not in recs_by_tag for t in need):
        return None
    fy, q1, q2, q3 = (recs_by_tag[t] for t in ["FY", "Q1", "Q2", "Q3"])
    flow = lambda k: fy[k] - q1[k] - q2[k] - q3[k]
    q4 = {"quarter": "Q4", "year": fy["year"], "is_annual": False, "derived": True}
    for k in ["total_revenue", "cogs", "net_profit_total"]:
        q4[k] = flow(k)
    # งบดุล Q4: ต้นงวด = ปลาย Q3 (ก.ย.), ปลายงวด = ปลายปี (จากไฟล์ FY)
    for base in ["trade_receivables", "inventory", "trade_payables",
                 "total_assets", "total_equity"]:
        q4[base + "_begin"] = q3[base + "_end"]
        q4[base + "_end"] = fy[base + "_end"]
    return q4


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    if argv[1] == "--dir":
        pats = ["*.XLS", "*.xls", "*.XLSX", "*.xlsx", "*.XLSM", "*.xlsm"]
        files = sorted({f for p in pats for f in glob.glob(os.path.join(argv[2], p))})
    else:
        files = argv[1:]
    if not files:
        print("ไม่พบไฟล์")
        return 1

    recs = []
    for f in files:
        try:
            rec = extract_file(f)
            recs.append(rec)
            unit = {1e6: "บาท", 1e3: "พันบาท", 1.0: "ล้านบาท"}.get(rec["pl_unit_div"], "?")
            tag = "FY" if rec["is_annual"] else rec["quarter"]
            print(f"✓ {rec['file']:32s} {tag}/{rec['year']}  หน่วย={unit}  "
                  f"balance_ok={rec.get('balance_ok')}")
        except ExtractError as e:
            print(f"✗ {e}")
            return 2

    print("\n=== ค่าที่ดึงได้ (ล้านบาท) ===")
    hdr = f"{'งวด':>7} {'รายได้':>12} {'ต้นทุนขาย':>12} {'กำไรสุทธิ':>12} {'AR_end':>10} {'INV_end':>10} {'AP_end':>10} {'สินทรัพย์':>12} {'ทุน':>12}"
    print(hdr)
    for r in recs:
        tag = "FY" if r["is_annual"] else r["quarter"]
        print(f"{tag+'/'+str(r['year']):>7} {r['total_revenue']:>12.1f} {r['cogs']:>12.1f} "
              f"{r['net_profit_total']:>12.1f} {r['trade_receivables_end']:>10.1f} "
              f"{r['inventory_end']:>10.1f} {r['trade_payables_end']:>10.1f} "
              f"{r['total_assets_end']:>12.1f} {r['total_equity_end']:>12.1f}")

    # ประกอบ Q4 จาก FY − Q1−Q2−Q3 (ถ้ามีครบ)
    by_tag = {("FY" if r["is_annual"] else r["quarter"]): r for r in recs}
    q4 = derive_q4(by_tag)
    if q4:
        rev4 = q4["total_revenue"]
        print(f"\n[Q4 = FY−Q1−Q2−Q3]  รายได้={rev4:.1f}  ต้นทุนขาย={q4['cogs']:.1f}  "
              f"กำไรสุทธิ={q4['net_profit_total']:.1f} ล้านบาท")

    print("\n=== 7 ตัวชี้วัดรายไตรมาส ===")
    print(f"{'งวด':>7} {'NPM':>8} {'ROA':>8} {'ROE':>8} {'DSO':>7} {'DIO':>7} {'DPO':>7} {'CCC':>7}")
    ordered = [by_tag[t] for t in ["Q1", "Q2", "Q3"] if t in by_tag]
    if q4:
        ordered.append(q4)
    for r in ordered:
        m = compute_quarter(r)
        print(f"{r['quarter']+'/'+str(r['year']):>7} {fmt(m['NPM'],1):>8} {fmt(m['ROA'],1):>8} "
              f"{fmt(m['ROE'],1):>8} {fmt(m['DSO']):>7} {fmt(m['DIO']):>7} "
              f"{fmt(m['DPO']):>7} {fmt(m['CCC']):>7}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
